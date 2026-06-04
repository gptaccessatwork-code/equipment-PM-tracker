"""
Production PM Tracker
====================
A manufacturing Preventive Maintenance (PM) tracker application.
- Multi-component tracking per machine (up to 5 components)
- Background email notification system
- System tray minimization
- Dark-themed industrial interface
- SQLite database storage
- Dynamic health color coding

Version: 1.0
Made by Sankar
"""

import sys
import os
import sqlite3
import json
import threading
import time
import smtplib
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import List, Dict, Optional

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QDialog, QDialogButtonBox,
    QScrollArea, QFrame, QGridLayout, QMessageBox, QSystemTrayIcon,
    QMenu, QProgressBar, QSpinBox, QDoubleSpinBox, QCheckBox,
    QComboBox, QGroupBox, QSplitter, QSpacerItem, QSizePolicy, QAbstractSpinBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QTextEdit, QTabBar, QInputDialog,
    QFileDialog
)
from PySide6.QtCore import Qt, QTimer, QSize, Signal, QObject, QEvent, QPoint
from PySide6.QtWidgets import QDateEdit
from PySide6.QtGui import QIcon, QPixmap, QColor, QFont, QPalette, QAction, QPainter, QPen, QBrush, QPolygon

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  THEME COLORS (Light Theme from Vercel UI)
# ══════════════════════════════════════════════════════════════════════════════
class Theme:
    # Background colors (Dark theme)
    BG_PRIMARY = "#0f172a"      # Dark slate
    BG_CARD = "#1e293b"          # Card background
    BG_MUTED = "#334155"         # Muted background
    BG_INPUT = "#0f172a"         # Input background
    
    # Border colors
    BORDER = "#334155"           # Border color
    BORDER_FOCUS = "#3b82f6"     # Focus border
    
    # Text colors
    TEXT_PRIMARY = "#f1f5f9"     # Primary text
    TEXT_MUTED = "#94a3b8"       # Muted text
    TEXT_DISABLED = "#64748b"    # Disabled text
    
    # Accent colors
    PRIMARY = "#3b82f6"          # Primary blue
    PRIMARY_HOVER = "#2563eb"    # Primary hover
    PRIMARY_LIGHT = "rgba(59, 130, 246, 0.1)"  # Primary light
    
    # Status colors
    GREEN = "#10b981"            # Healthy
    GREEN_LIGHT = "rgba(16, 185, 129, 0.1)"
    YELLOW = "#f59e0b"           # Warning
    YELLOW_LIGHT = "rgba(245, 158, 11, 0.1)"
    RED = "#ef4444"              # Critical
    RED_LIGHT = "rgba(239, 68, 68, 0.1)"
    
    # UI dimensions
    CORNER_RADIUS = 16
    BUTTON_HEIGHT = 44
    INPUT_HEIGHT = 44

class NoWheelDateEdit(QDateEdit):
    """Date edit that blocks mouse wheel changes."""

    def wheelEvent(self, event):
        event.ignore()

# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE SETUP
# ══════════════════════════════════════════════════════════════════════════════
def get_base_path():
    """Get the base directory for database and config files."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_db_path():
    """Get the database file path."""
    return os.path.join(get_base_path(), "production_pm_tracker.db")

def get_db_connection():
    """Open a SQLite connection with foreign keys enabled."""
    conn = sqlite3.connect(get_db_path())
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def get_email_config_path():
    """Get the email config file path."""
    return os.path.join(get_base_path(), "email_config.json")

def init_database():
    """Initialize the SQLite database with the required schema."""
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Create sheets table so each tab can stay isolated
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        try:
            cursor.execute("ALTER TABLE sheets ADD COLUMN email_recipients TEXT NOT NULL DEFAULT '[]'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("UPDATE sheets SET email_recipients = COALESCE(email_recipients, '[]')")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute(
                "INSERT OR IGNORE INTO sheets (id, name, email_recipients, created_at) VALUES (1, 'Default', ?, CURRENT_TIMESTAMP)",
                (json.dumps([]),)
            )
        except sqlite3.OperationalError:
            cursor.execute("""
                INSERT OR IGNORE INTO sheets (id, name, created_at)
                VALUES (1, 'Default', CURRENT_TIMESTAMP)
            """)
        
        # Create machines table without UNIQUE constraint on name
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS machines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                serial_number TEXT,
                location TEXT,
                sheet_id INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """) 
        
        # Migration: Remove UNIQUE constraint from existing databases
        try:
            # Check if the old table with UNIQUE constraint exists
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='machines'")
            table_sql = cursor.fetchone()
            if table_sql and 'UNIQUE' in table_sql[0]:
                # Create new table without UNIQUE constraint
                cursor.execute("""
                    CREATE TABLE machines_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        serial_number TEXT,
                        location TEXT,
                        sheet_id INTEGER NOT NULL DEFAULT 1,
                        created_at TEXT DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Copy data from old table to new table
                cursor.execute("""
                    INSERT INTO machines_new (id, name, serial_number, location, sheet_id, created_at)
                    SELECT id, name, serial_number, location, COALESCE(sheet_id, 1), created_at FROM machines
                """)
                
                # Drop old table
                cursor.execute("DROP TABLE machines")
                
                # Rename new table to original name
                cursor.execute("ALTER TABLE machines_new RENAME TO machines")
        except Exception:
            pass
        
        # Add new columns to existing machines table if they don't exist
        try:
            cursor.execute("ALTER TABLE machines ADD COLUMN serial_number TEXT")
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        try:
            cursor.execute("ALTER TABLE machines ADD COLUMN location TEXT")
        except sqlite3.OperationalError:
            pass  # Column already exists 

        try:
            cursor.execute("ALTER TABLE machines ADD COLUMN sheet_id INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("UPDATE machines SET sheet_id = COALESCE(sheet_id, 1)")
        except sqlite3.OperationalError:
            pass
        
        # Create components table with a real equipment tag
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS components (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                machine_id INTEGER,
                parent_machine_name TEXT NOT NULL,
                component_name TEXT NOT NULL,
                pm_interval_days INTEGER NOT NULL DEFAULT 30,
                alert_threshold_days INTEGER NOT NULL DEFAULT 5,
                last_performed_date TEXT,
                next_due_date TEXT,
                custom_start_date TEXT,
                custom_start_applied_date TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (machine_id) REFERENCES machines(id) ON DELETE CASCADE
            )
        """)

        # Migrate older component tables that were keyed only by machine name
        try:
            cursor.execute("PRAGMA table_info(components)")
            existing_columns = {row[1] for row in cursor.fetchall()}
            if existing_columns and "machine_id" not in existing_columns:
                cursor.execute("ALTER TABLE components RENAME TO components_old")
                cursor.execute("""
                    CREATE TABLE components (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        machine_id INTEGER,
                        parent_machine_name TEXT NOT NULL,
                        component_name TEXT NOT NULL,
                        pm_interval_days INTEGER NOT NULL DEFAULT 30,
                        alert_threshold_days INTEGER NOT NULL DEFAULT 5,
                        last_performed_date TEXT,
                        next_due_date TEXT,
                        custom_start_date TEXT,
                        custom_start_applied_date TEXT,
                        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (machine_id) REFERENCES machines(id) ON DELETE CASCADE
                    )
                """)
                cursor.execute("""
                    INSERT INTO components (
                        id, machine_id, parent_machine_name, component_name,
                        pm_interval_days, alert_threshold_days, last_performed_date,
                        next_due_date, custom_start_date, custom_start_applied_date, created_at
                    )
                    SELECT
                        c.id,
                        m.id,
                        c.parent_machine_name,
                        c.component_name,
                        c.pm_interval_days,
                        c.alert_threshold_days,
                        c.last_performed_date,
                        c.next_due_date,
                        c.custom_start_date,
                        date(c.created_at),
                        c.created_at
                    FROM components_old c
                    LEFT JOIN machines m ON m.name = c.parent_machine_name
                """)
                cursor.execute("DROP TABLE components_old")
        except Exception:
            pass

        # Add missing columns for existing databases
        try:
            cursor.execute("ALTER TABLE components ADD COLUMN machine_id INTEGER")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE components ADD COLUMN custom_start_date TEXT")
        except sqlite3.OperationalError:
            pass  # Column already exists

        try:
            cursor.execute("ALTER TABLE components ADD COLUMN custom_start_applied_date TEXT")
        except sqlite3.OperationalError:
            pass  # Column already exists

        try:
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_components_machine_id
                ON components(machine_id)
            """)
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("""
                UPDATE components
                SET custom_start_applied_date = date(created_at)
                WHERE custom_start_date IS NOT NULL
                  AND custom_start_applied_date IS NULL
            """)
        except sqlite3.OperationalError:
            pass

        # Create maintenance log table for component service history
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS maintenance_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                machine_id INTEGER NOT NULL,
                component_id INTEGER NOT NULL,
                sheet_id INTEGER NOT NULL DEFAULT 1,
                machine_name TEXT NOT NULL,
                component_name TEXT NOT NULL,
                maintenance_date TEXT NOT NULL,
                recorded_at TEXT NOT NULL,
                pm_interval_days INTEGER NOT NULL,
                next_due_date TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (machine_id) REFERENCES machines(id) ON DELETE CASCADE,
                FOREIGN KEY (component_id) REFERENCES components(id) ON DELETE CASCADE
            )
        """)

        try:
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_maintenance_log_component_date
                ON maintenance_log(component_id, maintenance_date DESC)
            """)
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_maintenance_log_machine_date
                ON maintenance_log(machine_id, maintenance_date DESC)
            """)
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE maintenance_log ADD COLUMN sheet_id INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("UPDATE maintenance_log SET sheet_id = COALESCE(sheet_id, 1)")
        except sqlite3.OperationalError:
            pass
        
        # Create email_log table to track one reminder per day
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS email_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_id INTEGER NOT NULL DEFAULT 1,
                sent_date TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                components_count INTEGER NOT NULL DEFAULT 0,
                recipients TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'sending',
                error_message TEXT
            )
        """)

        # Keep one log row per date so another app launch cannot send again
        try:
            cursor.execute("""
                DELETE FROM email_log
                WHERE id NOT IN (
                    SELECT MAX(id)
                    FROM email_log
                    GROUP BY sent_date, sheet_id
                )
            """)
            cursor.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_email_log_sent_date
                ON email_log(sent_date, sheet_id)
            """)
        except sqlite3.OperationalError:
            pass

        # Add new columns to older databases if they are missing
        try:
            cursor.execute("ALTER TABLE email_log ADD COLUMN status TEXT NOT NULL DEFAULT 'sending'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE email_log ADD COLUMN error_message TEXT")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE email_log ADD COLUMN sheet_id INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("UPDATE email_log SET sheet_id = COALESCE(sheet_id, 1)")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("""
                UPDATE sheets
                SET email_recipients = ?
                WHERE (email_recipients IS NULL OR TRIM(email_recipients) = '')
            """, (json.dumps([]),))
        except Exception:
            pass
        
        conn.commit()

# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE OPERATIONS
# ══════════════════════════════════════════════════════════════════════════════
class DatabaseManager:
    """Handle all database operations."""
    
    @staticmethod
    def get_sheets() -> List[Dict]:
        """Get all sheet tabs."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM sheets ORDER BY id")
                sheets = []
                for row in cursor.fetchall():
                    sheet = dict(row)
                    raw_recipients = sheet.get("email_recipients") or "[]"
                    try:
                        parsed = json.loads(raw_recipients)
                        if not isinstance(parsed, list):
                            parsed = []
                    except Exception:
                        parsed = [
                            item.strip()
                            for item in str(raw_recipients).replace("\n", ",").split(",")
                            if item.strip()
                        ]
                    sheet["email_recipients"] = parsed
                    sheet["to_addrs"] = parsed
                    sheets.append(sheet)
                return sheets
        except Exception as e:
            print(f"Error getting sheets: {e}")
            return []

    @staticmethod
    def get_sheet_by_id(sheet_id: int) -> Optional[Dict]:
        """Get a single sheet by its id."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM sheets WHERE id = ?", (sheet_id,))
                row = cursor.fetchone()
                if not row:
                    return None
                sheet = dict(row)
                raw_recipients = sheet.get("email_recipients") or "[]"
                try:
                    parsed = json.loads(raw_recipients)
                    if not isinstance(parsed, list):
                        parsed = []
                except Exception:
                    parsed = [
                        item.strip()
                        for item in str(raw_recipients).replace("\n", ",").split(",")
                        if item.strip()
                    ]
                sheet["email_recipients"] = parsed
                sheet["to_addrs"] = parsed
                return sheet
        except Exception as e:
            print(f"Error getting sheet {sheet_id}: {e}")
            return None

    @staticmethod
    def add_sheet(name: str, email_recipients: Optional[List[str]] = None) -> Optional[int]:
        """Add a new sheet tab."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                if email_recipients is None:
                    email_recipients = []
                cursor.execute(
                    "INSERT INTO sheets (name, email_recipients) VALUES (?, ?)",
                    (name.strip(), json.dumps(email_recipients))
                )
                conn.commit()
                return cursor.lastrowid
        except Exception as e:
            print(f"Error adding sheet: {e}")
            return None

    @staticmethod
    def update_sheet_recipients(sheet_id: int, email_recipients: List[str]) -> bool:
        """Update a sheet's recipient list."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE sheets SET email_recipients = ? WHERE id = ?",
                    (json.dumps(email_recipients), sheet_id)
                )
                conn.commit()
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error updating sheet recipients: {e}")
            return False

    @staticmethod
    def update_sheet_name(sheet_id: int, new_name: str) -> bool:
        """Rename a sheet tab."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE sheets SET name = ? WHERE id = ?",
                    (new_name.strip(), sheet_id)
                )
                conn.commit()
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error updating sheet name: {e}")
            return False

    @staticmethod
    def delete_sheet(sheet_id: int) -> bool:
        """Delete a sheet and all data attached to it."""
        if sheet_id == 1:
            return False

        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()

                cursor.execute("DELETE FROM email_log WHERE sheet_id = ?", (sheet_id,))
                cursor.execute("DELETE FROM maintenance_log WHERE sheet_id = ?", (sheet_id,))
                cursor.execute("SELECT id FROM machines WHERE sheet_id = ?", (sheet_id,))
                machine_ids = [row[0] for row in cursor.fetchall()]
                for machine_id in machine_ids:
                    cursor.execute("DELETE FROM machines WHERE id = ?", (machine_id,))
                cursor.execute("DELETE FROM sheets WHERE id = ?", (sheet_id,))

                conn.commit()
                return True
        except Exception as e:
            print(f"Error deleting sheet: {e}")
            return False

    @staticmethod
    def add_machine(name: str, serial_number: str = None, location: str = None, sheet_id: int = 1) -> Optional[int]:
        """Add a new machine and return its database id."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO machines (name, serial_number, location, sheet_id) VALUES (?, ?, ?, ?)",
                    (name, serial_number, location, sheet_id)
                )
                conn.commit()
                return cursor.lastrowid
        except Exception as e:
            print(f"Error adding machine: {e}")
            return None
    
    @staticmethod
    def add_component(machine_id: Optional[int], machine_name: str, component_data: Dict) -> bool:
        """Add a new component to a machine."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                if machine_id is None:
                    cursor.execute("SELECT id FROM machines WHERE name = ? ORDER BY id DESC LIMIT 1", (machine_name,))
                    result = cursor.fetchone()
                    machine_id = result[0] if result else None
                if machine_id is None:
                    raise ValueError(f"Unable to resolve machine id for '{machine_name}'")
                
                # Custom start date uses a one-time phase before the regular interval begins.
                last_performed = component_data.get('last_performed_date')
                next_due = None
                custom_start = component_data.get('custom_start_date')
                custom_start_applied = component_data.get('custom_start_applied_date')

                if custom_start:
                    try:
                        next_due = date.fromisoformat(custom_start).isoformat()
                    except ValueError:
                        next_due = custom_start
                    last_performed = None
                    if not custom_start_applied:
                        custom_start_applied = date.today().isoformat()
                elif last_performed:
                    try:
                        last_date = date.fromisoformat(last_performed)
                        interval = component_data.get('pm_interval_days', 30)
                        next_date = last_date + timedelta(days=interval)
                        next_due = next_date.isoformat()
                    except ValueError:
                        pass
                cursor.execute("""
                    INSERT INTO components 
                    (machine_id, parent_machine_name, component_name, pm_interval_days, 
                     alert_threshold_days, last_performed_date, next_due_date, custom_start_date, custom_start_applied_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    machine_id,
                    machine_name,
                    component_data['component_name'],
                    component_data['pm_interval_days'],
                    component_data['alert_threshold_days'],
                    last_performed,
                    next_due,
                    custom_start,
                    custom_start_applied
                ))
                conn.commit()
                return True
        except Exception as e:
            print(f"Error adding component: {e}")
            return False

    @staticmethod
    def _apply_component_schedule_state(cursor, component_id: int) -> bool:
        """Rebuild the current component schedule from its maintenance history."""
        try:
            cursor.execute("""
                SELECT pm_interval_days
                FROM components
                WHERE id = ?
            """, (component_id,))
            component_row = cursor.fetchone()
            if not component_row:
                return False

            interval_days = component_row["pm_interval_days"]
            cursor.execute("""
                SELECT maintenance_date
                FROM maintenance_log
                WHERE component_id = ?
                ORDER BY maintenance_date DESC, recorded_at DESC, id DESC
                LIMIT 1
            """, (component_id,))
            latest_log = cursor.fetchone()

            if latest_log:
                performed_date = latest_log["maintenance_date"]
                next_due_date = (date.fromisoformat(performed_date) + timedelta(days=interval_days)).isoformat()
            else:
                performed_date = None
                next_due_date = None

            cursor.execute("""
                UPDATE components
                SET last_performed_date = ?,
                    next_due_date = ?,
                    custom_start_date = NULL,
                    custom_start_applied_date = NULL
                WHERE id = ?
            """, (performed_date, next_due_date, component_id))
            return True
        except Exception as e:
            print(f"Error applying component schedule state: {e}")
            return False

    @staticmethod
    def update_component(component_id: int, machine_id: int, machine_name: str, component_data: Dict) -> bool:
        """Update an existing component in place."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()

                last_performed = component_data.get('last_performed_date')
                next_due = None
                custom_start = component_data.get('custom_start_date')
                custom_start_applied = component_data.get('custom_start_applied_date')

                if custom_start:
                    try:
                        next_due = date.fromisoformat(custom_start).isoformat()
                    except ValueError:
                        next_due = custom_start
                    last_performed = None
                    if not custom_start_applied:
                        custom_start_applied = date.today().isoformat()
                elif last_performed:
                    try:
                        last_date = date.fromisoformat(last_performed)
                        interval = component_data.get('pm_interval_days', 30)
                        next_date = last_date + timedelta(days=interval)
                        next_due = next_date.isoformat()
                    except ValueError:
                        pass

                cursor.execute("""
                    UPDATE components
                    SET machine_id = ?,
                        parent_machine_name = ?,
                        component_name = ?,
                        pm_interval_days = ?,
                        alert_threshold_days = ?,
                        last_performed_date = ?,
                        next_due_date = ?,
                        custom_start_date = ?,
                        custom_start_applied_date = ?
                    WHERE id = ?
                """, (
                    machine_id,
                    machine_name,
                    component_data['component_name'],
                    component_data['pm_interval_days'],
                    component_data['alert_threshold_days'],
                    last_performed,
                    next_due,
                    custom_start,
                    custom_start_applied,
                    component_id
                ))
                conn.commit()
                return True
        except Exception as e:
            print(f"Error updating component: {e}")
            return False

    @staticmethod
    def get_components_for_machine(machine_id: int) -> List[Dict]:
        """Get all components for a machine, ordered by id."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT *
                    FROM components
                    WHERE machine_id = ?
                    ORDER BY component_name, id
                """, (machine_id,))
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error getting machine components: {e}")
            return []

    @staticmethod
    def get_component_details(component_id: int) -> Optional[Dict]:
        """Get a single component with its owning machine details."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT
                        c.*,
                        COALESCE(m.name, c.parent_machine_name) AS machine_name,
                        m.serial_number AS machine_serial_number,
                        m.location AS machine_location,
                        m.sheet_id AS sheet_id
                    FROM components c
                    LEFT JOIN machines m ON c.machine_id = m.id
                    WHERE c.id = ?
                """, (component_id,))
                row = cursor.fetchone()
                return dict(row) if row else None
        except Exception as e:
            print(f"Error getting component details: {e}")
            return None
    
    @staticmethod
    def get_all_machines(sheet_id: Optional[int] = None) -> List[Dict]:
        """Get all machines with their components."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                if sheet_id is None:
                    cursor.execute("SELECT * FROM machines ORDER BY name")
                else:
                    cursor.execute("SELECT * FROM machines WHERE sheet_id = ? ORDER BY name", (sheet_id,))
                machines = []
                
                for row in cursor.fetchall():
                    machine = dict(row)
                    
                    # Get components for this machine
                    cursor.execute("""
                        SELECT * FROM components 
                        WHERE machine_id = ? OR (machine_id IS NULL AND parent_machine_name = ?)
                        ORDER BY component_name
                    """, (machine['id'], machine['name']))
                    
                    components = []
                    for comp_row in cursor.fetchall():
                        comp = dict(comp_row)
                        # Calculate days remaining
                        comp['days_remaining'] = DatabaseManager.calculate_days_remaining(
                            comp['next_due_date']
                        )
                        components.append(comp)
                    
                    machine['components'] = components
                    machines.append(machine)
                
                return machines
        except Exception as e:
            print(f"Error getting machines: {e}")
            return []
    
    @staticmethod
    def calculate_days_remaining(next_due_date: Optional[str]) -> int:
        """Calculate days remaining until next due date."""
        if not next_due_date:
            return 0
        try:
            due_date = date.fromisoformat(next_due_date)
            remaining = (due_date - date.today()).days
            return max(0, remaining)
        except ValueError:
            return 0
    
    @staticmethod
    def reset_component(component_id: int) -> bool:
        """Reset a component's maintenance to today."""
        return DatabaseManager.record_component_maintenance(component_id)

    @staticmethod
    def record_component_maintenance(
        component_id: int,
        maintenance_date: Optional[str] = None,
        notes: Optional[str] = None
    ) -> bool:
        """Record maintenance for a component, log it, and reset the interval."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # Get component and machine details first
                cursor.execute("""
                    SELECT
                        c.*,
                        COALESCE(m.name, c.parent_machine_name) AS machine_name,
                        m.serial_number AS machine_serial_number,
                        m.sheet_id AS sheet_id
                    FROM components c
                    LEFT JOIN machines m ON c.machine_id = m.id
                    WHERE c.id = ?
                """, (component_id,))
                result = cursor.fetchone()
                
                if not result:
                    return False
                
                component = dict(result)
                interval_days = component['pm_interval_days']
                performed_date = maintenance_date or date.today().isoformat()

                try:
                    performed_day = date.fromisoformat(performed_date)
                except ValueError:
                    return False

                next_due = (performed_day + timedelta(days=interval_days)).isoformat()
                
                # Update the component and append a maintenance history row
                cursor.execute("""
                    UPDATE components 
                    SET last_performed_date = ?, next_due_date = ?, custom_start_date = NULL, custom_start_applied_date = NULL
                    WHERE id = ?
                """, (performed_date, next_due, component_id))

                cursor.execute("""
                    INSERT INTO maintenance_log (
                        machine_id, component_id, sheet_id, machine_name, component_name,
                        maintenance_date, recorded_at, pm_interval_days, next_due_date, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    component['machine_id'],
                    component_id,
                    component.get('sheet_id', 1),
                    component['machine_name'],
                    component['component_name'],
                    performed_date,
                    datetime.now().isoformat(),
                    interval_days,
                    next_due,
                    notes
                ))
                if not DatabaseManager._apply_component_schedule_state(cursor, component_id):
                    conn.rollback()
                    return False
                conn.commit()
                return True
        except Exception as e:
            print(f"Error recording component maintenance: {e}")
            return False

    @staticmethod
    def update_maintenance_log(log_id: int, maintenance_date: str, notes: Optional[str] = None, sheet_id: int = 1) -> bool:
        """Edit a maintenance log entry and recalculate the component schedule."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT component_id
                    FROM maintenance_log
                    WHERE id = ? AND sheet_id = ?
                """, (log_id, sheet_id))
                log_row = cursor.fetchone()
                if not log_row:
                    return False

                component_id = log_row["component_id"]
                cursor.execute("""
                    SELECT pm_interval_days
                    FROM components
                    WHERE id = ?
                """, (component_id,))
                component_row = cursor.fetchone()
                if not component_row:
                    return False

                try:
                    performed_day = date.fromisoformat(maintenance_date)
                except ValueError:
                    return False

                interval_days = component_row["pm_interval_days"]
                next_due = (performed_day + timedelta(days=interval_days)).isoformat()
                cursor.execute("""
                    UPDATE maintenance_log
                    SET maintenance_date = ?,
                        recorded_at = ?,
                        pm_interval_days = ?,
                        next_due_date = ?,
                        notes = ?
                    WHERE id = ?
                """, (
                    maintenance_date,
                    datetime.now().isoformat(),
                    interval_days,
                    next_due,
                    notes,
                    log_id
                ))
                if not DatabaseManager._apply_component_schedule_state(cursor, component_id):
                    conn.rollback()
                    return False
                conn.commit()
                return True
        except Exception as e:
            print(f"Error updating maintenance log: {e}")
            return False

    @staticmethod
    def delete_maintenance_log(log_id: int, sheet_id: int = 1) -> bool:
        """Delete a maintenance log entry and recalculate the component schedule."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT component_id
                    FROM maintenance_log
                    WHERE id = ? AND sheet_id = ?
                """, (log_id, sheet_id))
                log_row = cursor.fetchone()
                if not log_row:
                    return False

                component_id = log_row["component_id"]
                cursor.execute("DELETE FROM maintenance_log WHERE id = ? AND sheet_id = ?", (log_id, sheet_id))
                if cursor.rowcount <= 0:
                    return False
                if not DatabaseManager._apply_component_schedule_state(cursor, component_id):
                    conn.rollback()
                    return False
                conn.commit()
                return True
        except Exception as e:
            print(f"Error deleting maintenance log: {e}")
            return False
    
    @staticmethod
    def delete_component(component_id: int) -> bool:
        """Delete a specific component."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM components WHERE id = ?", (component_id,))
                conn.commit()
                return True
        except Exception as e:
            print(f"Error deleting component: {e}")
            return False
    
    @staticmethod
    def delete_machine(machine_id: int) -> bool:
        """Delete a machine and all its components (CASCADE will handle components)."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM machines WHERE id = ?", (machine_id,))
                conn.commit()
                return True
        except Exception as e:
            print(f"Error deleting machine: {e}")
            return False
    
    @staticmethod
    def was_email_sent_today(sent_date: str, sheet_id: int = 1) -> bool:
        """Check if today's reminder has already been claimed or sent for a sheet."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT 1
                    FROM email_log
                    WHERE sent_date = ? AND sheet_id = ?
                    LIMIT 1
                """, (sent_date, sheet_id))
                return cursor.fetchone() is not None
        except Exception as e:
            print(f"Error checking email log: {e}")
            return False

    @staticmethod
    def claim_daily_email(sent_date: str, sheet_id: int = 1) -> bool:
        """Claim today's reminder slot before sending to prevent duplicates."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO email_log (
                        sheet_id, sent_date, sent_at, components_count, recipients, status, error_message
                    )
                    VALUES (?, ?, ?, 0, '', 'sending', NULL)
                """, (sheet_id, sent_date, datetime.now().isoformat()))
                conn.commit()
                return True
        except sqlite3.IntegrityError:
            return False
        except Exception as e:
            print(f"Error claiming email slot: {e}")
            return False

    @staticmethod
    def mark_email_sent(sent_date: str, components_count: int, recipients: str, sheet_id: int = 1) -> bool:
        """Update the daily reminder row after a successful send."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE email_log
                    SET sent_at = ?, components_count = ?, recipients = ?, status = 'sent', error_message = NULL
                    WHERE sent_date = ? AND sheet_id = ?
                """, (datetime.now().isoformat(), components_count, recipients, sent_date, sheet_id))
                conn.commit()
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error updating email log: {e}")
            return False

    @staticmethod
    def mark_email_failed(sent_date: str, error_message: str, sheet_id: int = 1) -> bool:
        """Record a failed send attempt for today's reminder."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE email_log
                    SET status = 'failed', error_message = ?
                    WHERE sent_date = ? AND sheet_id = ?
                """, (error_message, sent_date, sheet_id))
                conn.commit()
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error updating failed email log: {e}")
            return False
    
    @staticmethod
    def update_machine_details(machine_id: int, old_name: str, new_name: str, serial_number: str = None, location: str = None) -> bool:
        """Update a machine's details including name, serial number, and location."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Update machine name if changed
                if old_name != new_name:
                    # Update machine name
                    cursor.execute("""
                        UPDATE machines SET name = ? WHERE id = ?
                    """, (new_name, machine_id))
                    
                    # Update component parent references
                    cursor.execute("""
                        UPDATE components SET parent_machine_name = ? WHERE machine_id = ?
                    """, (new_name, machine_id))
                
                # Update serial number and location
                cursor.execute("""
                    UPDATE machines SET serial_number = ?, location = ? WHERE id = ?
                """, (serial_number, location, machine_id))
                
                conn.commit()
                return True
        except Exception as e:
            print(f"Error updating machine details: {e}")
            return False
    
    @staticmethod
    def update_machine_name(machine_id: int, new_name: str) -> bool:
        """Update a machine's name."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Update machine name
                cursor.execute("""
                    UPDATE machines SET name = ? WHERE id = ?
                """, (new_name, machine_id))
                
                # Update component parent references
                cursor.execute("""
                    UPDATE components SET parent_machine_name = ? WHERE machine_id = ?
                """, (new_name, machine_id))
                
                conn.commit()
                return True
        except Exception as e:
            print(f"Error updating machine name: {e}")
            return False
    
    @staticmethod
    def delete_components_by_machine_id(machine_id: int) -> bool:
        """Delete all components for a machine."""
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    DELETE FROM components WHERE machine_id = ?
                """, (machine_id,))
                conn.commit()
                return True
        except Exception as e:
            print(f"Error deleting components: {e}")
            return False
    
    @staticmethod
    def get_components_due_soon(sheet_id: Optional[int] = None) -> List[Dict]:
        """Get components that are due soon or overdue."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                query = """
                    SELECT
                        c.*,
                        COALESCE(m.name, c.parent_machine_name) as machine_name,
                        m.serial_number as machine_serial_number,
                        m.sheet_id as sheet_id
                    FROM components c
                    LEFT JOIN machines m ON c.machine_id = m.id
                    WHERE c.next_due_date IS NOT NULL
                    ORDER BY c.next_due_date ASC
                """
                params = []
                if sheet_id is not None:
                    query = query.replace("WHERE c.next_due_date IS NOT NULL", "WHERE c.next_due_date IS NOT NULL AND m.sheet_id = ?")
                    params.append(sheet_id)

                cursor.execute(query, params)
                
                due_components = []
                for row in cursor.fetchall():
                    comp = dict(row)
                    days_remaining = DatabaseManager.calculate_days_remaining(
                        comp['next_due_date']
                    )
                    
                    # Check if due within alert threshold or overdue
                    if days_remaining <= comp['alert_threshold_days']:
                        comp['days_remaining'] = days_remaining
                        due_components.append(comp)
                
                return due_components
        except Exception as e:
            print(f"Error getting due components: {e}")
            return []

    @staticmethod
    def get_maintenance_history(sheet_id: Optional[int] = None, machine_id: Optional[int] = None, component_id: Optional[int] = None) -> List[Dict]:
        """Get maintenance log entries for a machine or component."""
        try:
            with get_db_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                query = """
                    SELECT *
                    FROM maintenance_log
                """
                params = []
                filters = []
                if machine_id is not None:
                    filters.append("machine_id = ?")
                    params.append(machine_id)
                if component_id is not None:
                    filters.append("component_id = ?")
                    params.append(component_id)
                if sheet_id is not None:
                    filters.append("sheet_id = ?")
                    params.append(sheet_id)
                if filters:
                    query += " WHERE " + " AND ".join(filters)
                query += " ORDER BY maintenance_date DESC, recorded_at DESC, id DESC"

                cursor.execute(query, params)
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error getting maintenance history: {e}")
            return []

# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
class EmailConfig:
    """Handle email configuration management."""
    
    DEFAULT_CONFIG = {
        "smtp_host": "smtp.gmail.com",
        "smtp_port": 587,
        "smtp_user": "",
        "smtp_password": "",
        "from_addr": "",
        "to_addrs": [],
        "enabled": False
    }
    
    @staticmethod
    def load_config() -> Dict:
        """Load email configuration from file."""
        try:
            with open(get_email_config_path(), 'r') as f:
                config = json.load(f)
                # Merge with defaults to ensure all keys exist
                merged = EmailConfig.DEFAULT_CONFIG.copy()
                merged.update(config)
                return merged
        except FileNotFoundError:
            return EmailConfig.DEFAULT_CONFIG.copy()
        except Exception as e:
            print(f"Error loading email config: {e}")
            return EmailConfig.DEFAULT_CONFIG.copy()
    
    @staticmethod
    def save_config(config: Dict) -> bool:
        """Save email configuration to file."""
        try:
            with open(get_email_config_path(), 'w') as f:
                json.dump(config, f, indent=2)
            return True
        except Exception as e:
            print(f"Error saving email config: {e}")
            return False
    
    @staticmethod
    def send_email(config: Dict, subject: str, html_body: str) -> tuple[bool, str]:
        """Send email using SMTP."""
        try:
            if not config.get('enabled') or not config.get('to_addrs'):
                return False, "Email not enabled or no recipients"
            
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = config["from_addr"]
            msg["To"] = ", ".join(config["to_addrs"])
            msg.attach(MIMEText(html_body, "html"))
            
            with smtplib.SMTP(
                config["smtp_host"], 
                int(config["smtp_port"]), 
                timeout=15
            ) as server:
                server.starttls()
                server.login(config["smtp_user"], config["smtp_password"])
                server.sendmail(
                    config["from_addr"], 
                    config["to_addrs"], 
                    msg.as_string()
                )
            
            return True, ""
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    def send_test_email(config: Dict) -> tuple[bool, str]:
        """Send a test email with [TEST] prefix in subject."""
        # Create test component data
        test_components = [
            {
                'component_name': 'TEST-COMPONENT',
                'machine_name': 'TEST-MACHINE',
                'days_remaining': 5
            }
        ]
        
        subject, html = EmailConfig.build_alert_email(test_components, sheet_name="Default")
        # Add [TEST] prefix to subject
        subject = "[TEST] " + subject
        
        return EmailConfig.send_email(config, subject, html)
    
    @staticmethod
    def build_alert_email(components: List[Dict], sheet_name: str = "Default") -> tuple[str, str]:
        """Build INFICON-style grouped HTML email content."""
        today_str = datetime.now().strftime("%d %b %Y")
        total = len(components)
        subject = f"[PM Tracker] {sheet_name}: Maintenance Due Within 30 Days — {today_str}"
        
        def _status_label(days):
            if days <= 0: return "#fca5a5", "OVERDUE"
            if days < 5: return "#fcd34d", f"{days} days remaining"
            return "#6ee7b7", f"{days} days remaining"
        
        # Group components by machine
        machines = {}
        for comp in components:
            machine_name = comp.get('machine_name', 'Unknown')
            if machine_name not in machines:
                machines[machine_name] = []
            machines[machine_name].append(comp)
        
        # Build HTML sections for each machine
        sections_html = ""
        for machine_name, machine_components in machines.items():
            machine_serial_number = next(
                (
                    item.get('machine_serial_number')
                    for item in machine_components
                    if item.get('machine_serial_number')
                ),
                None
            )
            machine_title = machine_name
            if machine_serial_number:
                machine_title = f"{machine_name} (SN: {machine_serial_number})"

            machine_components_sorted = sorted(machine_components, key=lambda x: x['days_remaining'])
            
            # Build rows for this machine
            rows = ""
            for idx, item in enumerate(machine_components_sorted):
                col, status_txt = _status_label(item['days_remaining'])
                bg = "#ffffff" if idx % 2 == 0 else "#f7faf9"
                
                rows += (
                    f'<tr>'
                    f'<td style="padding:8px 12px;border-bottom:1px solid #e8e8e8;'
                    f'font-weight:bold;color:#1c1917;background:{bg};">'
                    f'{item["component_name"]}</td>'
                    f'<td style="padding:8px 12px;border-bottom:1px solid #e8e8e8;'
                    f'font-weight:bold;color:{col};background:{bg};">{status_txt}</td>'
                    f'</tr>'
                )
            
            overdue_count = sum(1 for x in machine_components if x['days_remaining'] <= 0)
            intro = (
                f"The following {len(machine_components)} component(s) for "
                f"<strong>{machine_title}</strong> are due within 30 days or are already overdue."
            )
            if overdue_count:
                intro += f" <strong style='color:#fca5a5;'>{overdue_count} component(s) are already overdue.</strong>"
            
            sections_html += f"""
            <!-- ═══ {machine_name} ═══ -->
            <tr><td style="padding:0 32px 24px;">
              <table width="100%" cellpadding="0" cellspacing="0">
                <tr>
                  <td style="background:#5fd1c8;border-radius:6px 6px 0 0;padding:14px 16px;">
                    <div style="font-size:15px;font-weight:700;color:#fff;">{machine_title}</div>
                    <div style="font-size:12px;color:#a0c0e8;margin-top:3px;">
                      {len(machine_components)} component(s) require attention
                    </div>
                  </td>
                </tr>
                <tr><td style="padding:12px 16px 4px;background:#fafbfc;">
                  <p style="margin:0;color:#444;font-size:13px;">{intro}</p>
                </td></tr>
                <tr><td>
                  <table width="100%" cellpadding="0" cellspacing="0"
                         style="border-collapse:collapse;font-size:13px;">
                    <thead>
                      <tr style="background:#5fd1c8;">
                        <th style="padding:10px 12px;text-align:center;color:#fff;
                               border-bottom:2px solid rgba(255,255,255,.3);">Component</th>
                        <th style="padding:10px 12px;text-align:center;color:#fff;
                               border-bottom:2px solid rgba(255,255,255,.3);">Status</th>
                      </tr>
                    </thead>
                    <tbody>{rows}</tbody>
                  </table>
                </td></tr>
              </table>
            </td></tr>
            <tr><td style="padding:0 32px 8px;"><hr style="border:none;border-top:1px solid #eee;"></td></tr>
            """
        
        html = f"""<!DOCTYPE html>
<html>
<body style="font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;margin:0;padding:0;">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td style="padding:30px 0;text-align:center;">
<table width="700" cellpadding="0" cellspacing="0"
       style="margin:0 auto;background:#fff;border-radius:8px;
              box-shadow:0 2px 8px rgba(0,0,0,.08);">

  <!-- Header -->
  <tr>
    <td style="background:#5fd1c8;border-radius:8px 8px 0 0;padding:28px 32px;">
      <div style="font-size:20px;font-weight:700;color:#fff;">Equipment Maintenance Alert - {sheet_name}</div>
      <div style="font-size:13px;color:#a0c0e8;margin-top:6px;">
        {today_str} &nbsp;|&nbsp; {total} component(s) require attention
        across {len(machines)} equipment
      </div>
    </td>
  </tr>

  <!-- Intro -->
  <tr>
    <td style="padding:24px 32px 16px;">
      <p style="margin:0;color:#444;font-size:14px;line-height:1.6;">
        This is your scheduled maintenance reminder.
        Please schedule service at the earliest opportunity.
      </p>
    </td>
  </tr>

  <!-- Sections (one per equipment) -->
  {sections_html}

  <!-- Footer -->
  <tr>
    <td style="padding:16px 32px;background:#f0f4f8;border-radius:0 0 8px 8px;
               font-size:11px;color:#999;border-top:1px solid #e0e0e0;">
      Automated notification — Equipment PM Tracker
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>"""
        return subject, html

# ══════════════════════════════════════════════════════════════════════════════
#  BACKGROUND EMAIL NOTIFICATION THREAD
# ══════════════════════════════════════════════════════════════════════════════
class EmailNotificationThread(threading.Thread):
    """Background thread for email notifications."""
    
    CHECK_INTERVAL = 3600  # 1 hour
    
    def __init__(self):
        super().__init__(daemon=True)
        self._stop_event = threading.Event()
        self._callback = None  # Callback to update UI
    
    def set_callback(self, callback):
        """Set callback for UI updates."""
        self._callback = callback
    
    def stop(self):
        """Stop the notification thread."""
        self._stop_event.set()
    
    def run(self):
        """Run the notification loop."""
        time.sleep(15)  # Initial delay
        
        while not self._stop_event.is_set():
            self._check_notifications()
            self._stop_event.wait(self.CHECK_INTERVAL)
    
    def _check_notifications(self):
        """Check for components due soon and send emails."""
        config = EmailConfig.load_config()
        
        if not config.get('enabled'):
            return
        
        today = date.today().isoformat()
        sheets = DatabaseManager.get_sheets()
        sent_count = 0
        failed_count = 0

        for sheet in sheets:
            sheet_id = sheet.get('id', 1)
            sheet_name = sheet.get('name', 'Default')

            if DatabaseManager.was_email_sent_today(today, sheet_id):
                continue

            due_components = DatabaseManager.get_components_due_soon(sheet_id)
            if not due_components:
                continue

            sheet_recipients = sheet.get("email_recipients") or []
            if not sheet_recipients:
                continue

            if not DatabaseManager.claim_daily_email(today, sheet_id):
                continue

            subject, html = EmailConfig.build_alert_email(due_components, sheet_name=sheet_name)
            send_config = config.copy()
            send_config["to_addrs"] = sheet_recipients
            success, error = EmailConfig.send_email(send_config, subject, html)

            if success:
                DatabaseManager.mark_email_sent(
                    today,
                    len(due_components),
                    ", ".join(config.get('to_addrs', [])),
                    sheet_id
                )
                sent_count += 1
            else:
                DatabaseManager.mark_email_failed(today, error, sheet_id)
                failed_count += 1

        if self._callback and (sent_count or failed_count):
            self._callback(f"Email run complete: {sent_count} sheet(s) sent, {failed_count} failed")

# ══════════════════════════════════════════════════════════════════════════════
#  STYLED WIDGETS
# ══════════════════════════════════════════════════════════════════════════════
class StyledButton(QPushButton):
    """Styled button with dark theme."""
    
    def __init__(self, text: str, primary: bool = True, parent=None):
        super().__init__(text, parent)
        self.primary = primary
        self._apply_style()
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setMinimumHeight(Theme.BUTTON_HEIGHT)
    
    def _apply_style(self):
        if self.primary:
            self.setStyleSheet(f"""
                QPushButton {{
                    background-color: {Theme.PRIMARY};
                    color: white;
                    border: none;
                    border-radius: {Theme.CORNER_RADIUS}px;
                    padding: 12px 24px;
                    font-weight: 600;
                    font-size: 14px;
                }}
                QPushButton:hover {{
                    background-color: {Theme.PRIMARY_HOVER};
                }}
                QPushButton:pressed {{
                    background-color: #3db8a8;
                }}
                QPushButton:disabled {{
                    background-color: {Theme.TEXT_DISABLED};
                    color: {Theme.TEXT_MUTED};
                }}
            """)
        else:
            self.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    color: {Theme.TEXT_PRIMARY};
                    border: 2px solid {Theme.BORDER};
                    border-radius: {Theme.CORNER_RADIUS}px;
                    padding: 12px 24px;
                    font-weight: 600;
                    font-size: 14px;
                }}
                QPushButton:hover {{
                    border-color: {Theme.PRIMARY};
                    color: {Theme.PRIMARY};
                }}
                QPushButton:pressed {{
                    background-color: {Theme.BORDER};
                }}
                QPushButton:disabled {{
                    color: {Theme.TEXT_DISABLED};
                    border-color: {Theme.TEXT_DISABLED};
                }}
            """)

class StyledLineEdit(QLineEdit):
    """Styled line edit with dark theme."""
    
    def __init__(self, placeholder: str = "", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self.setMinimumHeight(Theme.INPUT_HEIGHT)
        self._apply_style()
    
    def _apply_style(self):
        self.setStyleSheet(f"""
            QLineEdit {{
                background-color: {Theme.BG_INPUT};
                color: {Theme.TEXT_PRIMARY};
                border: 2px solid {Theme.BORDER};
                border-radius: 12px;
                padding: 12px 16px;
                font-size: 14px;
            }}
            QLineEdit:focus {{
                border-color: {Theme.BORDER_FOCUS};
            }}
            QLineEdit:disabled {{
                color: {Theme.TEXT_DISABLED};
            }}
        """)

class StyledSpinBox(QSpinBox):
    """Styled spin box with dark theme."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(Theme.INPUT_HEIGHT)
        self.setMinimum(1)
        self.setMaximum(3650)  # 10 years
        self.setValue(30)
        self.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        self._apply_style()
    
    def _apply_style(self):
        self.setStyleSheet(f"""
            QSpinBox {{
                background-color: {Theme.BG_INPUT};
                color: {Theme.TEXT_PRIMARY};
                border: 2px solid {Theme.BORDER};
                border-radius: 12px;
                padding: 12px 16px;
                font-size: 14px;
            }}
            QSpinBox:focus {{
                border-color: {Theme.BORDER_FOCUS};
            }}
            QSpinBox::up-button, QSpinBox::down-button {{
                background-color: transparent;
                border: none;
                width: 20px;
            }}
            QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
                background-color: {Theme.PRIMARY};
            }}
        """)
    
    def wheelEvent(self, event):
        """Disable mouse wheel scrolling to prevent accidental value changes."""
        event.ignore()

class StyledCard(QFrame):
    """Styled card with dark theme."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._apply_style()
    
    def _apply_style(self):
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {Theme.BG_CARD};
                border: 1px solid {Theme.BORDER};
                border-radius: {Theme.CORNER_RADIUS}px;
            }}
        """)
        self.setProperty("class", "card")

class StyledProgressBar(QProgressBar):
    """Styled progress bar with health colors."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTextVisible(False)
        self.setMaximumHeight(8)
        self._apply_style()
    
    def _apply_style(self):
        self.setStyleSheet(f"""
            QProgressBar {{
                background-color: {Theme.BG_MUTED};
                border: none;
                border-radius: 4px;
            }}
            QProgressBar::chunk {{
                background-color: {Theme.GREEN};
                border-radius: 4px;
            }}
        """)
    
    def set_health_color(self, days_remaining: int, alert_threshold: int):
        """Update color based on health status."""
        if days_remaining <= 0:
            color = Theme.RED
        elif days_remaining <= alert_threshold:
            color = Theme.YELLOW
        else:
            color = Theme.GREEN
        
        self.setStyleSheet(f"""
            QProgressBar {{
                background-color: {Theme.BG_MUTED};
                border: none;
                border-radius: 4px;
            }}
            QProgressBar::chunk {{
                background-color: {color};
                border-radius: 4px;
            }}
        """)

class MetricCard(StyledCard):
    """Clickable metric card used for dashboard filters."""

    clicked = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._active = False
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)

    def set_active(self, active: bool):
        self._active = active
        self.setProperty("active", active)
        self.style().unpolish(self)
        self.style().polish(self)
        self.update()

class ComponentRow(QWidget):
    """Row widget for displaying component status."""
    
    reset_clicked = Signal(int)  # Signal with component ID
    delete_clicked = Signal(int)  # Signal with component ID
    
    def __init__(self, component_data: Dict, parent=None):
        super().__init__(parent)
        self.component_data = component_data
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QHBoxLayout()
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(12)
        layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        # Component name
        name_label = QLabel(self.component_data['component_name'])
        name_label.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                font-weight: 600;
                font-size: 14px;
                min-width: 160px;
                border: none;
            }}
        """)
        layout.addWidget(name_label)
        
        # Progress bar
        self.progress_bar = StyledProgressBar()
        days_remaining = self.component_data.get('days_remaining', 0)
        total_days = self._get_progress_total_days(days_remaining)
        alert_threshold = self.component_data.get('alert_threshold_days', 5)
        
        self.progress_bar.setRange(0, total_days)
        self.progress_bar.setValue(min(days_remaining, total_days))
        self.progress_bar.set_health_color(days_remaining, alert_threshold)
        layout.addWidget(self.progress_bar, 1)
        
        # Days remaining badge
        days_text = f"{days_remaining}d left" if days_remaining > 0 else "OVERDUE"
        days_label = QLabel(days_text)
        
        if days_remaining <= 0:
            bg_color = Theme.RED_LIGHT
            text_color = Theme.RED
        elif days_remaining <= alert_threshold:
            bg_color = Theme.YELLOW_LIGHT
            text_color = Theme.YELLOW
        else:
            bg_color = Theme.GREEN_LIGHT
            text_color = Theme.GREEN
        
        days_label.setStyleSheet(f"""
            QLabel {{
                background-color: {bg_color};
                color: {text_color};
                padding: 4px 12px;
                border: none;
                border-radius: 12px;
                font-weight: 600;
                font-size: 12px;
                min-width: 70px;
                text-align: center;
            }}
        """)
        layout.addWidget(days_label)
        
        # Record maintenance button
        reset_btn = QPushButton("✓")
        reset_btn.setFixedSize(32, 32)
        reset_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background: transparent;
                color: {Theme.TEXT_MUTED};
                font-size: 16px;
            }}
            QPushButton:hover {{
                color: {Theme.GREEN};
            }}
        """)
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_btn.clicked.connect(lambda: self.reset_clicked.emit(self.component_data['id']))
        reset_btn.setToolTip(f"Record maintenance for {self.component_data['component_name']}")
        
        # Delete button
        delete_btn = QPushButton("✕")
        delete_btn.setFixedSize(32, 32)
        delete_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background: transparent;
                color: {Theme.TEXT_MUTED};
                font-size: 16px;
            }}
            QPushButton:hover {{
                color: {Theme.RED};
            }}
        """)
        delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_btn.clicked.connect(lambda: self.delete_clicked.emit(self.component_data['id']))
        delete_btn.setToolTip(f"Delete {self.component_data['component_name']}")
        
        # Action buttons layout with tight spacing
        action_layout = QHBoxLayout()
        action_layout.setSpacing(2)  # Reduced spacing between buttons
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.addWidget(reset_btn)
        action_layout.addWidget(delete_btn)
        
        layout.addLayout(action_layout)
        
        self.setLayout(layout)
        
        # Add hover effect
        self.setStyleSheet(f"""
            QWidget {{
                background-color: transparent;
                border: none;
                border-radius: 12px;
            }}
            QWidget:hover {{
                background-color: {Theme.BG_MUTED};
            }}
        """)

    def _get_progress_total_days(self, days_remaining: int) -> int:
        """Return the active phase length for the progress bar."""
        custom_start_date = self.component_data.get('custom_start_date')
        custom_start_applied_date = self.component_data.get('custom_start_applied_date')
        last_performed_date = self.component_data.get('last_performed_date')
        next_due_date = self.component_data.get('next_due_date')

        # Before the first PM, the custom date is the active due date.
        if (
            custom_start_date
            and next_due_date == custom_start_date
        ):
            phase_start = custom_start_applied_date
            if not phase_start:
                phase_start = self.component_data.get('created_at')

            try:
                if phase_start:
                    phase_start_date = date.fromisoformat(str(phase_start).split(" ")[0])
                    phase_end_date = date.fromisoformat(custom_start_date)
                    return max((phase_end_date - phase_start_date).days, 1)
            except ValueError:
                pass

        return max(self.component_data.get('pm_interval_days', 30), 1)

class RecordMaintenanceDialog(QDialog):
    """Dialog for recording a completed maintenance action."""

    def __init__(
        self,
        machine_name: str,
        component_name: str,
        parent=None,
        maintenance_date: Optional[str] = None,
        notes: Optional[str] = None
    ):
        super().__init__(parent)
        self.machine_name = machine_name
        self.component_name = component_name
        self._initial_maintenance_date = maintenance_date
        self._initial_notes = notes
        self._setup_ui()

    def _setup_ui(self):
        self.setWindowTitle("Edit Maintenance" if self._initial_maintenance_date or self._initial_notes else "Record Maintenance")
        self.setMinimumWidth(460)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
        """)

        layout = QVBoxLayout()
        layout.setSpacing(14)

        title = QLabel("Edit Maintenance" if self._initial_maintenance_date or self._initial_notes else "Record Maintenance")
        title.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                font-size: 18px;
                font-weight: 700;
                border: none;
            }}
        """)
        layout.addWidget(title)

        summary = QLabel(f"{self.machine_name}  •  {self.component_name}")
        summary.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_MUTED};
                font-size: 13px;
                border: none;
            }}
        """)
        layout.addWidget(summary)

        date_label = QLabel("Maintenance Date")
        date_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; border: none; }}")
        layout.addWidget(date_label)

        self.date_input = NoWheelDateEdit()
        self.date_input.setCalendarPopup(True)
        if self._initial_maintenance_date:
            try:
                self.date_input.setDate(date.fromisoformat(self._initial_maintenance_date))
            except ValueError:
                self.date_input.setDate(date.today())
        else:
            self.date_input.setDate(date.today())
        self.date_input.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.date_input.setStyleSheet(f"""
            QDateEdit {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 8px;
            }}
            QDateEdit:focus {{
                border: 1px solid {Theme.BORDER_FOCUS};
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background-color: {Theme.BG_CARD};
                padding: 4px 8px;
                min-height: 32px;
            }}
            QCalendarWidget QToolButton {{
                color: {Theme.TEXT_PRIMARY};
                background-color: transparent;
                min-width: 28px;
                min-height: 28px;
                padding: 4px;
                margin: 2px;
                border-radius: 4px;
                font-size: 14px;
            }}
            QCalendarWidget QToolButton#qt_calendar_prevmonth,
            QCalendarWidget QToolButton#qt_calendar_nextmonth {{
                min-width: 28px;
                max-width: 28px;
                min-height: 28px;
                max-height: 28px;
                padding: 2px;
                qproperty-iconSize: 16px 16px;
            }}
            QCalendarWidget QToolButton#qt_calendar_monthbutton,
            QCalendarWidget QToolButton#qt_calendar_yearbutton {{
                padding: 4px 16px 4px 8px;
            }}
            QCalendarWidget QToolButton#qt_calendar_monthbutton::menu-indicator {{
                subcontrol-position: center right;
                subcontrol-origin: padding;
                right: 2px;
                width: 10px;
                height: 10px;
            }}
            QCalendarWidget QToolButton:hover {{
                background-color: {Theme.BG_MUTED};
            }}
            QCalendarWidget QMenu {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
            }}
            QCalendarWidget QSpinBox {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                selection-background-color: {Theme.PRIMARY};
                selection-color: {Theme.TEXT_PRIMARY};
            }}
            QCalendarWidget QAbstractItemView:enabled {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_PRIMARY};
                selection-background-color: {Theme.PRIMARY};
                selection-color: {Theme.TEXT_PRIMARY};
            }}
            QCalendarWidget QAbstractItemView:disabled {{
                color: {Theme.TEXT_DISABLED};
            }}
        """)
        layout.addWidget(self.date_input)

        notes_label = QLabel("Notes (optional)")
        notes_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; border: none; }}")
        layout.addWidget(notes_label)

        self.notes_input = QTextEdit()
        self.notes_input.setPlaceholderText("Add any useful notes about what was completed...")
        self.notes_input.setMinimumHeight(110)
        self.notes_input.setStyleSheet(f"""
            QTextEdit {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 8px;
                padding: 8px;
            }}
            QTextEdit:focus {{
                border: 1px solid {Theme.BORDER_FOCUS};
            }}
        """)
        if self._initial_notes:
            self.notes_input.setPlainText(self._initial_notes)
        layout.addWidget(self.notes_input)

        button_layout = QHBoxLayout()
        button_layout.addStretch()

        cancel_btn = StyledButton("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        save_btn = StyledButton("Save", primary=True)
        save_btn.clicked.connect(self.accept)
        button_layout.addWidget(save_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def get_data(self) -> tuple[str, str]:
        """Return the selected maintenance date and notes."""
        maintenance_date = self.date_input.date().toPython().isoformat()
        notes = self.notes_input.toPlainText().strip() or None
        return maintenance_date, notes

class MaintenanceHistoryDialog(QDialog):
    """Dialog for reviewing maintenance history for an equipment item."""

    def __init__(self, machine_data: Dict, sheet_id: int = 1, parent=None):
        super().__init__(parent)
        self.machine_data = machine_data
        self.sheet_id = sheet_id
        self._all_logs = []
        self._visible_logs = []
        self._setup_ui()
        self._load_logs()

    def _setup_ui(self):
        self.setWindowTitle("Maintenance History")
        self.setMinimumSize(760, 480)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
        """)

        layout = QVBoxLayout()
        layout.setSpacing(14)

        title = QLabel(f"Maintenance History - {self.machine_data['name']}")
        title.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                font-size: 18px;
                font-weight: 700;
                border: none;
            }}
        """)
        layout.addWidget(title)

        details = []
        serial_number = self.machine_data.get('serial_number')
        location = self.machine_data.get('location')
        if serial_number:
            details.append(f"SN: {serial_number}")
        if location:
            details.append(f"Location: {location}")
        detail_label = QLabel("  •  ".join(details) if details else "Maintenance entries recorded for this equipment.")
        detail_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 13px; border: none; }}")
        layout.addWidget(detail_label)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(8)
        filter_label = QLabel("Component")
        filter_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; border: none; }}")
        filter_layout.addWidget(filter_label)

        self.component_filter = QComboBox()
        self.component_filter.setStyleSheet(f"""
            QComboBox {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 10px;
                min-width: 220px;
            }}
        """)
        self.component_filter.currentIndexChanged.connect(self._refresh_table)
        filter_layout.addWidget(self.component_filter)

        sort_label = QLabel("Sort")
        sort_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; border: none; margin-left: 12px; }}")
        filter_layout.addWidget(sort_label)

        self.sort_order = QComboBox()
        self.sort_order.addItem("Newest first", "desc")
        self.sort_order.addItem("Oldest first", "asc")
        self.sort_order.setStyleSheet(f"""
            QComboBox {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 10px;
                min-width: 150px;
            }}
        """)
        self.sort_order.currentIndexChanged.connect(self._refresh_table)
        filter_layout.addWidget(self.sort_order)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Date", "Component", "Notes"])
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.itemSelectionChanged.connect(self._update_action_state)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                gridline-color: {Theme.BORDER};
                border: 1px solid {Theme.BORDER};
                border-radius: 8px;
            }}
            QHeaderView::section {{
                background-color: {Theme.BG_MUTED};
                color: {Theme.TEXT_PRIMARY};
                padding: 8px;
                border: none;
                font-weight: 600;
            }}
        """)
        header = self.table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table, 1)

        action_layout = QHBoxLayout()
        action_layout.addStretch()

        self.edit_btn = StyledButton("Edit Selected", primary=False)
        self.edit_btn.clicked.connect(self._edit_selected_log)
        action_layout.addWidget(self.edit_btn)

        self.delete_btn = StyledButton("Delete Selected", primary=False)
        self.delete_btn.clicked.connect(self._delete_selected_log)
        action_layout.addWidget(self.delete_btn)

        close_btn = StyledButton("Close", primary=False)
        close_btn.clicked.connect(self.reject)
        action_layout.addWidget(close_btn)
        layout.addLayout(action_layout)

        self.setLayout(layout)

    def _load_logs(self):
        self._all_logs = DatabaseManager.get_maintenance_history(sheet_id=self.sheet_id, machine_id=self.machine_data['id'])
        self.component_filter.blockSignals(True)
        self.component_filter.clear()
        self.component_filter.addItem("All Components", None)
        component_names = []
        for component in self.machine_data.get('components', []):
            component_name = component.get('component_name')
            if component_name and component_name not in component_names:
                component_names.append(component_name)
        for component_name in sorted(component_names):
            self.component_filter.addItem(component_name, component_name)
        self.component_filter.blockSignals(False)
        self._refresh_table()

    def _refresh_table(self):
        selected_component = self.component_filter.currentData()
        logs = self._all_logs
        if selected_component:
            logs = [log for log in logs if log.get('component_name') == selected_component]

        sort_order = self.sort_order.currentData() or "desc"
        logs = sorted(
            logs,
            key=lambda log: (
                log.get('maintenance_date', ''),
                log.get('recorded_at', ''),
                log.get('id', 0)
            ),
            reverse=(sort_order == "desc")
        )

        self._visible_logs = logs

        self.table.clearSpans()
        self.table.setRowCount(0)
        if not logs:
            self._update_action_state()
            self.table.setRowCount(1)
            empty_item = QTableWidgetItem("No maintenance history yet.")
            empty_item.setForeground(QColor(Theme.TEXT_MUTED))
            self.table.setItem(0, 0, empty_item)
            self.table.setSpan(0, 0, 1, 3)
            return

        self.table.setRowCount(len(logs))
        for row_index, entry in enumerate(logs):
            values = [
                entry.get('maintenance_date', ''),
                entry.get('component_name', ''),
                entry.get('notes') or ''
            ]
            for col_index, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_index in (0, 1):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_index, col_index, item)

        self._update_action_state()

    def _update_action_state(self):
        has_selection = self._selected_log() is not None
        self.edit_btn.setEnabled(has_selection)
        self.delete_btn.setEnabled(has_selection)

    def _selected_log(self) -> Optional[Dict]:
        selected_rows = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not selected_rows:
            return None
        row = selected_rows[0].row()
        if row < 0 or row >= len(self._visible_logs):
            return None
        return self._visible_logs[row]

    def _refresh_parent(self):
        parent = self.parent()
        if parent and hasattr(parent, "_refresh_data"):
            parent._refresh_data()

    def _edit_selected_log(self):
        log = self._selected_log()
        if not log:
            return

        dialog = RecordMaintenanceDialog(
            log.get('machine_name', self.machine_data['name']),
            log.get('component_name', 'Component'),
            self,
            maintenance_date=log.get('maintenance_date'),
            notes=log.get('notes')
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        maintenance_date, notes = dialog.get_data()
        if DatabaseManager.update_maintenance_log(log['id'], maintenance_date, notes, self.sheet_id):
            self._load_logs()
            self._refresh_parent()
        else:
            QMessageBox.warning(self, "Edit Failed", "Could not update the selected log entry.")

    def _delete_selected_log(self):
        log = self._selected_log()
        if not log:
            return

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Delete Log Entry")
        msg_box.setText(
            f"Delete the maintenance entry for {log.get('component_name', 'this component')} "
            f"on {log.get('maintenance_date', '')}?"
        )
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)

        if msg_box.exec() != QMessageBox.StandardButton.Yes:
            return

        if DatabaseManager.delete_maintenance_log(log['id'], self.sheet_id):
            self._load_logs()
            self._refresh_parent()
        else:
            QMessageBox.warning(self, "Delete Failed", "Could not delete the selected log entry.")

class SheetManagementDialog(QDialog):
    """Dialog for renaming and deleting sheet tabs."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self._load_sheets()

    def _setup_ui(self):
        self.setWindowTitle("Manage Sheets")
        self.setMinimumSize(620, 380)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
        """)

        layout = QVBoxLayout()
        layout.setSpacing(14)

        title = QLabel("Manage Sheets")
        title.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                font-size: 18px;
                font-weight: 700;
                border: none;
            }}
        """)
        layout.addWidget(title)

        hint = QLabel("Rename sheets or delete a sheet and everything inside it.")
        hint.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 13px; border: none; }}")
        layout.addWidget(hint)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["ID", "Sheet", "Items", "Recipients"])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.cellDoubleClicked.connect(self._handle_cell_double_click)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                gridline-color: {Theme.BORDER};
                border: 1px solid {Theme.BORDER};
                border-radius: 8px;
            }}
            QHeaderView::section {{
                background-color: {Theme.BG_MUTED};
                color: {Theme.TEXT_PRIMARY};
                padding: 8px;
                border: none;
                font-weight: 600;
            }}
        """)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table, 1)

        button_row = QHBoxLayout()
        button_row.addStretch()

        self.rename_btn = StyledButton("Rename", primary=False)
        self.rename_btn.clicked.connect(self._rename_selected)
        button_row.addWidget(self.rename_btn)

        self.delete_btn = StyledButton("Delete", primary=False)
        self.delete_btn.clicked.connect(self._delete_selected)
        button_row.addWidget(self.delete_btn)

        self.recipients_btn = StyledButton("Recipients", primary=False)
        self.recipients_btn.clicked.connect(self._edit_recipients)
        button_row.addWidget(self.recipients_btn)

        close_btn = StyledButton("Close", primary=False)
        close_btn.clicked.connect(self.reject)
        button_row.addWidget(close_btn)

        layout.addLayout(button_row)
        self.setLayout(layout)

    def _load_sheets(self):
        sheets = DatabaseManager.get_sheets()
        self.table.setRowCount(len(sheets))
        for row_index, sheet in enumerate(sheets):
            sheet_id = sheet.get("id", 1)
            item_count = self._count_items(sheet_id)
            recipients = sheet.get("email_recipients") or []
            recipients_text = f"{len(recipients)} recipient(s)" if recipients else "No recipients"

            id_item = QTableWidgetItem(str(sheet_id))
            id_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            name_item = QTableWidgetItem(sheet.get("name", "Sheet"))
            items_item = QTableWidgetItem(str(item_count))
            items_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            recipients_item = QTableWidgetItem(recipients_text)
            recipients_item.setData(Qt.ItemDataRole.UserRole, recipients)

            self.table.setItem(row_index, 0, id_item)
            self.table.setItem(row_index, 1, name_item)
            self.table.setItem(row_index, 2, items_item)
            self.table.setItem(row_index, 3, recipients_item)

    def _count_items(self, sheet_id: int) -> int:
        machines = DatabaseManager.get_all_machines(sheet_id)
        return sum(len(machine.get("components", [])) for machine in machines)

    def _selected_sheet(self) -> Optional[Dict]:
        selected_rows = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not selected_rows:
            return None
        row = selected_rows[0].row()
        if row < 0 or row >= self.table.rowCount():
            return None
        sheet_id_item = self.table.item(row, 0)
        name_item = self.table.item(row, 1)
        if not sheet_id_item or not name_item:
            return None
        try:
            sheet_id = int(sheet_id_item.text())
        except ValueError:
            return None
        recipients_item = self.table.item(row, 3)
        recipients_text = recipients_item.data(Qt.ItemDataRole.UserRole) if recipients_item else []
        if not isinstance(recipients_text, list):
            recipients_text = []
        return {
            "id": sheet_id,
            "name": name_item.text(),
            "email_recipients": recipients_text,
        }

    def _sheet_at_row(self, row: int) -> Optional[Dict]:
        if row < 0 or row >= self.table.rowCount():
            return None
        sheet_id_item = self.table.item(row, 0)
        name_item = self.table.item(row, 1)
        if not sheet_id_item or not name_item:
            return None
        try:
            sheet_id = int(sheet_id_item.text())
        except ValueError:
            return None
        recipients_item = self.table.item(row, 3)
        recipients_text = recipients_item.data(Qt.ItemDataRole.UserRole) if recipients_item else []
        if not isinstance(recipients_text, list):
            recipients_text = []
        return {
            "id": sheet_id,
            "name": name_item.text(),
            "email_recipients": recipients_text,
        }

    def _select_row(self, row: int):
        if row < 0 or row >= self.table.rowCount():
            return
        self.table.selectRow(row)

    def _handle_cell_double_click(self, row: int, column: int):
        if column == 1:
            self._select_row(row)
            self._rename_selected()

    def _show_context_menu(self, pos):
        index = self.table.indexAt(pos)
        if not index.isValid():
            return

        row = index.row()
        self._select_row(row)
        sheet = self._sheet_at_row(row)
        if not sheet:
            return

        menu = QMenu(self)
        rename_action = QAction("Rename", self)
        delete_action = QAction("Delete", self)
        rename_action.triggered.connect(self._rename_selected)
        delete_action.triggered.connect(self._delete_selected)
        menu.addAction(rename_action)
        menu.addAction(delete_action)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _rename_selected(self):
        sheet = self._selected_sheet()
        if not sheet:
            return

        new_name, accepted = QInputDialog.getText(
            self,
            "Rename Sheet",
            "Enter a new sheet name:",
            QLineEdit.EchoMode.Normal,
            sheet["name"]
        )
        if not accepted:
            return

        new_name = new_name.strip()
        if not new_name:
            QMessageBox.warning(self, "Invalid Name", "Please enter a sheet name.")
            return

        if new_name == sheet["name"]:
            return

        if not DatabaseManager.update_sheet_name(sheet["id"], new_name):
            QMessageBox.warning(self, "Rename Failed", "Could not rename the sheet.")
            return

        self._load_sheets()
        self._refresh_parent()

    def _edit_recipients(self):
        sheet = self._selected_sheet()
        if not sheet:
            return

        current = sheet.get("email_recipients") or []
        current_text = ", ".join(current)
        text, accepted = QInputDialog.getMultiLineText(
            self,
            "Sheet Recipients",
            "Enter comma-separated email recipients for this sheet:",
            current_text
        )
        if not accepted:
            return

        recipients = [item.strip() for item in text.replace("\n", ",").split(",") if item.strip()]
        if not DatabaseManager.update_sheet_recipients(sheet["id"], recipients):
            QMessageBox.warning(self, "Update Failed", "Could not update sheet recipients.")
            return

        self._load_sheets()
        self._refresh_parent()

    def _delete_selected(self):
        sheet = self._selected_sheet()
        if not sheet:
            return

        if sheet["id"] == 1:
            QMessageBox.information(self, "Default Sheet", "The Default sheet cannot be deleted.")
            return

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Delete Sheet")
        msg_box.setText(
            f"Delete '{sheet['name']}' and all equipment, components, maintenance logs, and reminder history inside it?"
        )
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        if msg_box.exec() != QMessageBox.StandardButton.Yes:
            return

        if DatabaseManager.delete_sheet(sheet["id"]):
            self._load_sheets()
            self._refresh_parent()
        else:
            QMessageBox.warning(self, "Delete Failed", "Could not delete the selected sheet.")

    def _refresh_parent(self):
        parent = self.parent()
        if parent and hasattr(parent, "_load_sheet_tabs"):
            parent._load_sheet_tabs()
class EquipmentCard(StyledCard):
    """Card widget for displaying equipment and its components."""
    
    reset_component = Signal(int)  # Signal with component ID
    edit_requested = Signal(dict)  # Signal with machine data for editing
    delete_requested = Signal(dict)  # Signal with machine data
    history_requested = Signal(dict)  # Signal with machine data
    refresh_needed = Signal()  # Signal to refresh the card
    
    def __init__(self, machine_data: Dict, parent=None):
        super().__init__(parent)
        self.machine_data = machine_data
        self.setFixedHeight(320)
        self._setup_ui()
        self._apply_hover_style()
        
        # Enable double-click for editing
        self.setCursor(Qt.CursorShape.PointingHandCursor)
    
    def _setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(8)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        # Machine name header with delete button
        header_layout = QHBoxLayout()
        header_layout.setSpacing(8)
        header_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        name_label = QLabel(self.machine_data['name'])
        name_label.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                font-weight: 700;
                font-size: 18px;
                border: none;
            }}
        """)
        header_layout.addWidget(name_label)
        
        header_layout.addStretch()

        # Maintenance history button
        history_btn = QPushButton("Logs")
        history_btn.setFixedSize(48, 28)
        history_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background: transparent;
                color: {Theme.TEXT_MUTED};
                font-size: 11px;
                font-weight: 700;
            }}
            QPushButton:hover {{
                color: {Theme.PRIMARY};
            }}
        """)
        history_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        history_btn.clicked.connect(lambda: self.history_requested.emit(self.machine_data))
        history_btn.setToolTip(f"View maintenance history for {self.machine_data['name']}")
        header_layout.addWidget(history_btn)
        
        # Delete equipment button
        delete_btn = QPushButton("✕")
        delete_btn.setFixedSize(28, 28)
        delete_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background: transparent;
                color: {Theme.TEXT_MUTED};
                font-size: 18px;
            }}
            QPushButton:hover {{
                color: {Theme.RED};
            }}
        """)
        delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_btn.clicked.connect(lambda: self.delete_requested.emit(self.machine_data))
        delete_btn.setToolTip(f"Delete {self.machine_data['name']}")
        header_layout.addWidget(delete_btn)
        
        layout.addLayout(header_layout)
        
        # Serial number and location info
        serial_number = self.machine_data.get('serial_number', '')
        location = self.machine_data.get('location', '')
        
        if serial_number or location:
            info_layout = QHBoxLayout()
            info_layout.setSpacing(12)
            info_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
            
            if serial_number:
                serial_label = QLabel(f"SN: {serial_number}")
                serial_label.setStyleSheet(f"""
                    QLabel {{
                        color: {Theme.TEXT_MUTED};
                        font-size: 12px;
                        border: none;
                    }}
                """)
                info_layout.addWidget(serial_label)
            
            if location:
                location_label = QLabel(f"📍 {location}")
                location_label.setStyleSheet(f"""
                    QLabel {{
                        color: {Theme.TEXT_MUTED};
                        font-size: 12px;
                        border: none;
                    }}
                """)
                info_layout.addWidget(location_label)
            
            layout.addLayout(info_layout)
        
        # Components list
        components_layout = QVBoxLayout()
        components_layout.setSpacing(8)
        components_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        components = self.machine_data.get('components', [])
        if components:
            for component in components:
                row = ComponentRow(component)
                row.setFixedHeight(45)
                row.reset_clicked.connect(self.reset_component.emit)
                row.delete_clicked.connect(self._delete_component)
                components_layout.addWidget(row)
        else:
            no_components = QLabel("No components added yet")
            no_components.setStyleSheet(f"""
                QLabel {{
                    color: {Theme.TEXT_MUTED};
                    font-style: italic;
                    text-align: center;
                    padding: 20px;
                    border: none;
                }}
            """)
            no_components.setAlignment(Qt.AlignmentFlag.AlignCenter)
            components_layout.addWidget(no_components)
        
        layout.addLayout(components_layout)
        
        # Add vertical spacer to push content to top
        spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
        layout.addSpacerItem(spacer)
        self.setLayout(layout)
    
    def _delete_component(self, component_id: int):
        """Handle component deletion with confirmation."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Delete Component")
        msg_box.setText("Are you sure you want to delete this component and its maintenance history?")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        # Style the box to match dark industrial theme
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        reply = msg_box.exec()
        
        if reply == QMessageBox.StandardButton.Yes:
            if DatabaseManager.delete_component(component_id):
                # Emit signal to refresh the card
                self.refresh_needed.emit()
    
    def _apply_hover_style(self):
        """Apply hover effect styling to the equipment card."""
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {Theme.BG_CARD};
                border: 1px solid {Theme.BORDER};
                border-radius: {Theme.CORNER_RADIUS}px;
            }}
            QFrame:hover {{
                border: 1px solid {Theme.GREEN};
            }}
        """)
    
    def mouseDoubleClickEvent(self, event):
        """Handle double-click event to open edit dialog."""
        self.edit_requested.emit(self.machine_data)
        super().mouseDoubleClickEvent(event)

# ══════════════════════════════════════════════════════════════════════════════
#  ADD EQUIPMENT DIALOG
# ══════════════════════════════════════════════════════════════════════════════
class ComponentInputWidget(QWidget):
    """Widget for inputting a single component's data."""
    
    delete_requested = Signal()
    
    def __init__(self, component_num: int, parent=None):
        super().__init__(parent)
        self.component_num = component_num
        self._setup_ui()
    
    def _setup_ui(self):
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        card = QFrame()
        card.setObjectName("componentInputCard")
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        card_layout = QGridLayout(card)
        card_layout.setContentsMargins(18, 14, 18, 14)
        card_layout.setHorizontalSpacing(14)
        card_layout.setVerticalSpacing(14)
        
        # Component number header with delete button
        header_layout = QHBoxLayout()
        header_layout.setSpacing(8)
        
        header = QLabel(f"Component {self.component_num}")
        header.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_MUTED};
                font-weight: 600;
                font-size: 12px;
                border: none;
                background-color: transparent;
            }}
        """)
        header_layout.addWidget(header)
        
        header_layout.addStretch()
        
        # Delete button
        delete_btn = QPushButton("✕")
        delete_btn.setFixedSize(24, 24)
        delete_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background: transparent;
                color: {Theme.TEXT_MUTED};
                font-size: 14px;
            }}
            QPushButton:hover {{
                color: {Theme.RED};
            }}
        """)
        delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_btn.clicked.connect(self.delete_requested.emit)
        delete_btn.setToolTip("Remove this component")
        header_layout.addWidget(delete_btn)
        
        card_layout.addLayout(header_layout, 0, 0, 1, 4)
        
        # Component name
        name_label = QLabel("Name:")
        name_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 12px; border: none; background-color: transparent; }}")
        card_layout.addWidget(name_label, 1, 0)
        
        self.name_input = StyledLineEdit("e.g., Calibration")
        self.name_input.setMaxLength(20)
        card_layout.addWidget(self.name_input, 1, 1, 1, 2)
        
        # Interval days
        interval_label = QLabel("Interval (days):")
        interval_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 12px; border: none; background-color: transparent; }}")
        card_layout.addWidget(interval_label, 2, 0)
        
        self.interval_input = StyledSpinBox()
        card_layout.addWidget(self.interval_input, 2, 1)
        
        # Alert threshold
        alert_label = QLabel("Alert (days):")
        alert_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 12px; border: none; background-color: transparent; }}")
        card_layout.addWidget(alert_label, 2, 2)
        
        self.alert_input = StyledSpinBox()
        self.alert_input.setValue(5)
        card_layout.addWidget(self.alert_input, 2, 3)
        
        # Custom start date option
        self.custom_start_checkbox = QCheckBox("Custom Start Date")
        self.custom_start_checkbox.setStyleSheet(f"""
            QCheckBox {{
                color: {Theme.TEXT_PRIMARY};
                font-size: 12px;
            }}
            QCheckBox::indicator {{
                width: 18px;
                height: 18px;
                border: 2px solid {Theme.TEXT_MUTED};
                border-radius: 3px;
                background-color: {Theme.BG_CARD};
            }}
            QCheckBox::indicator:hover {{
                border: 2px solid {Theme.PRIMARY};
            }}
            QCheckBox::indicator:checked {{
                background-color: {Theme.PRIMARY};
                border: 2px solid {Theme.PRIMARY};
            }}
        """)
        self.custom_start_date = NoWheelDateEdit()
        self.custom_start_date.setDate(date.today())
        self.custom_start_date.setCalendarPopup(True)
        self.custom_start_date.setVisible(False)
        self.custom_start_date.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.custom_start_date.setStyleSheet(f"""
            QDateEdit {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 8px;
            }}
            QDateEdit:focus {{
                border: 1px solid {Theme.BORDER_FOCUS};
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background-color: {Theme.BG_CARD};
                padding: 4px 8px;
                min-height: 32px;
            }}
            QCalendarWidget QToolButton {{
                color: {Theme.TEXT_PRIMARY};
                background-color: transparent;
                min-width: 28px;
                min-height: 28px;
                padding: 4px;
                margin: 2px;
                border-radius: 4px;
                font-size: 14px;
            }}
            QCalendarWidget QToolButton#qt_calendar_prevmonth,
            QCalendarWidget QToolButton#qt_calendar_nextmonth {{
                min-width: 28px;
                max-width: 28px;
                min-height: 28px;
                max-height: 28px;
                padding: 2px;
                qproperty-iconSize: 16px 16px;
            }}
            QCalendarWidget QToolButton#qt_calendar_monthbutton,
            QCalendarWidget QToolButton#qt_calendar_yearbutton {{
                padding: 4px 16px 4px 8px;
            }}
            QCalendarWidget QToolButton#qt_calendar_monthbutton::menu-indicator {{
                subcontrol-position: center right;
                subcontrol-origin: padding;
                right: 2px;
                width: 10px;
                height: 10px;
            }}
            QCalendarWidget QToolButton:hover {{
                background-color: {Theme.BG_MUTED};
            }}
            QCalendarWidget QMenu {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
            }}
            QCalendarWidget QSpinBox {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_INPUT};
                selection-background-color: {Theme.PRIMARY};
                selection-color: {Theme.TEXT_PRIMARY};
            }}
            QCalendarWidget QAbstractItemView:enabled {{
                color: {Theme.TEXT_PRIMARY};
                background-color: {Theme.BG_PRIMARY};
                selection-background-color: {Theme.PRIMARY};
                selection-color: {Theme.TEXT_PRIMARY};
            }}
            QCalendarWidget QAbstractItemView:disabled {{
                color: {Theme.TEXT_DISABLED};
            }}
        """)
        self.custom_start_checkbox.stateChanged.connect(self._on_custom_start_changed)
        
        card_layout.addWidget(self.custom_start_checkbox, 3, 0, 1, 4)
        card_layout.addWidget(self.custom_start_date, 4, 0, 1, 4)

        card.setStyleSheet(f"""
            QFrame#componentInputCard {{
                background-color: {Theme.BG_CARD};
                border: 1px solid rgba(148, 163, 184, 0.7);
                border-radius: 14px;
            }}
            QFrame#componentInputCard:hover {{
                border: 1px solid {Theme.PRIMARY};
            }}
        """)

        outer_layout.addWidget(card)
    
    def _on_custom_start_changed(self, state):
        """Handle custom start date checkbox state change."""
        self.custom_start_date.setVisible(state == 2)  # 2 = checked
    
    def get_data(self) -> Optional[Dict]:
        """Get component data from inputs."""
        name = self.name_input.text().strip()
        if not name:
            return None
        
        data = {
            'component_name': name,
            'pm_interval_days': self.interval_input.value(),
            'alert_threshold_days': self.alert_input.value(),
        }
        
        # Handle custom start date
        if self.custom_start_checkbox.isChecked():
            custom_date = self.custom_start_date.date().toPython()
            custom_date_str = custom_date.isoformat()
            data['custom_start_date'] = custom_date_str

            existing_custom_start_date = getattr(self, '_existing_custom_start_date', None)
            existing_custom_applied = getattr(self, '_existing_custom_start_applied_date', None)
            if existing_custom_start_date == custom_date_str and existing_custom_applied:
                data['custom_start_applied_date'] = existing_custom_applied
            else:
                data['custom_start_applied_date'] = date.today().isoformat()
        else:
            data['last_performed_date'] = date.today().isoformat()
        
        return data

class AddEquipmentDialog(QDialog):
    """Dialog for adding or editing equipment with components."""
    
    def __init__(self, machine_data: Optional[Dict] = None, sheet_id: int = 1, parent=None):
        super().__init__(parent)
        self.machine_data = machine_data
        self.sheet_id = machine_data.get('sheet_id', sheet_id) if machine_data else sheet_id
        self.is_edit_mode = machine_data is not None
        self.component_widgets = []
        self._setup_ui()
        if self.is_edit_mode:
            self._load_existing_data()
    
    def _setup_ui(self):
        title_text = "Edit Equipment" if self.is_edit_mode else "Add Equipment"
        self.setWindowTitle(title_text)
        self.resize(820, 780)
        self.setMinimumWidth(760)
        self.setMinimumHeight(920)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QScrollArea {{
                border: none;
                background-color: transparent;
                background: transparent;
            }}
            QScrollArea > QWidget > QWidget {{
                background-color: transparent;
                background: transparent;
            }}
            /* Modern Scrollbar Styling */
            QScrollBar:vertical {{
                border: none;
                background: transparent;
                width: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.TEXT_MUTED};
                min-height: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar:horizontal {{
                border: none;
                background: transparent;
                height: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:horizontal {{
                background: {Theme.TEXT_MUTED};
                min-width: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(20)
        
        # Title
        title = QLabel(title_text)
        title.setStyleSheet(f"""
            QLabel {{
                font-size: 20px;
                font-weight: 700;
                color: {Theme.TEXT_PRIMARY};
                border: none;
            }}
        """)
        layout.addWidget(title)
        
        description_text = "Edit equipment and maintenance schedules." if self.is_edit_mode else "Set up a new piece of equipment with maintenance schedules."
        description = QLabel(description_text)
        description.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 14px; border: none; }}")
        layout.addWidget(description)
        
        # Equipment name
        name_label = QLabel("Equipment Name:")
        name_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(name_label)
        
        self.equipment_name_input = StyledLineEdit("e.g., Torque Wrench (max 44 chars)")
        self.equipment_name_input.setMaxLength(44)  # Limit to 44 characters
        layout.addWidget(self.equipment_name_input)
        
        # Serial Number
        serial_label = QLabel("Serial Number (optional):")
        serial_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(serial_label)
        
        self.serial_input = StyledLineEdit("e.g., SN-12345")
        layout.addWidget(self.serial_input)
        
        # Location
        location_label = QLabel("Location (optional):")
        location_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(location_label)
        
        self.location_input = StyledLineEdit("e.g., Cell 1")
        layout.addWidget(self.location_input)
        
        # Components section
        components_label = QLabel("Components")
        components_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(components_label)
        
        self.component_count_label = QLabel("1 of 5")
        self.component_count_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 12px; border: none; }}")
        layout.addWidget(self.component_count_label)
        
        # Scroll area for components
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(460)
        scroll.setStyleSheet(f"""
            QScrollArea {{
                border: none;
                background-color: transparent;
            }}
            /* Modern Scrollbar Styling */
            QScrollBar:vertical {{
                border: none;
                background: transparent;
                width: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.TEXT_MUTED};
                min-height: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)
        
        self.components_container = QWidget()
        self.components_layout = QVBoxLayout()
        self.components_layout.setSpacing(14)
        self.components_container.setLayout(self.components_layout)
        self.components_container.setStyleSheet("""
            QWidget {
                background-color: transparent;
                background: transparent;
            }
        """)
        scroll.setWidget(self.components_container)
        scroll.viewport().setStyleSheet("""
            QWidget {
                background-color: transparent;
                background: transparent;
            }
        """)
        layout.addWidget(scroll)
        
        # Add first component only in add mode
        if not self.is_edit_mode:
            self._add_component()
        
        # Add component button
        self.add_component_btn = StyledButton("+ Add Component", primary=False)
        self.add_component_btn.clicked.connect(self._add_component)
        layout.addWidget(self.add_component_btn)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_btn = StyledButton("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        self.save_btn = StyledButton("Save", primary=True)
        self.save_btn.clicked.connect(self._save)
        self.save_btn.setEnabled(False)
        button_layout.addWidget(self.save_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        # Enable validation
        self.equipment_name_input.textChanged.connect(self._validate_inputs)
        
        # Initial validation
        self._validate_inputs()
    
    def _add_component(self):
        """Add a new component input widget."""
        if len(self.component_widgets) >= 5:
            return
        
        component_num = len(self.component_widgets) + 1
        widget = ComponentInputWidget(component_num)
        widget.delete_requested.connect(lambda: self._remove_component(widget))
        self.component_widgets.append(widget)
        self.components_layout.addWidget(widget)
        
        # Update count label
        self.component_count_label.setText(f"{len(self.component_widgets)} of 5")
        
        # Disable add button if limit reached
        if len(self.component_widgets) >= 5:
            self.add_component_btn.setEnabled(False)
            self.add_component_btn.setText("Maximum components reached")
        
        # Enable validation
        widget.name_input.textChanged.connect(self._validate_inputs)
        self._validate_inputs()
    
    def _remove_component(self, widget: ComponentInputWidget):
        """Remove a component input widget."""
        if widget in self.component_widgets:
            widget.deleteLater()
            self.component_widgets.remove(widget)
            
            # Renumber remaining components
            for i, w in enumerate(self.component_widgets):
                w.component_num = i + 1
                # Find and update the header label
                header = w.findChild(QLabel)
                if header:
                    header.setText(f"Component {i + 1}")
            
            # Update count label
            self.component_count_label.setText(f"{len(self.component_widgets)} of 5")
            
            # Re-enable add button if below limit
            if len(self.component_widgets) < 5:
                self.add_component_btn.setEnabled(True)
                self.add_component_btn.setText("+ Add Component")
            
            self._validate_inputs()
    
    def _validate_inputs(self):
        """Validate inputs and enable/disable save button."""
        equipment_name = self.equipment_name_input.text().strip()
        has_components = any(
            widget.name_input.text().strip() 
            for widget in self.component_widgets
        )
        
        # Only enable/disable save button if it exists (called during setup)
        if hasattr(self, 'save_btn'):
            self.save_btn.setEnabled(bool(equipment_name and has_components))
    
    def _load_existing_data(self):
        """Load existing machine data into the dialog for editing."""
        if not self.machine_data:
            return
        
        # Set equipment name
        self.equipment_name_input.setText(self.machine_data['name'])
        
        # Set serial number
        self.serial_input.setText(self.machine_data.get('serial_number', ''))
        
        # Set location
        self.location_input.setText(self.machine_data.get('location', ''))
        
        # Clear existing component widgets
        for widget in self.component_widgets:
            widget.deleteLater()
        self.component_widgets = []
        
        # Load components
        components = self.machine_data.get('components', [])
        for i, component in enumerate(components):
            widget = ComponentInputWidget(i + 1)
            widget.name_input.setText(component['component_name'])
            widget.interval_input.setValue(component['pm_interval_days'])
            widget.alert_input.setValue(component['alert_threshold_days'])
            custom_start = component.get('custom_start_date')
            if custom_start:
                widget.custom_start_checkbox.setChecked(True)
                try:
                    widget.custom_start_date.setDate(date.fromisoformat(custom_start))
                except ValueError:
                    pass
                widget._existing_custom_start_date = custom_start
                widget._existing_custom_start_applied_date = component.get('custom_start_applied_date')
            widget.delete_requested.connect(lambda: self._remove_component(widget))
            self.component_widgets.append(widget)
            self.components_layout.addWidget(widget)
        
        # Update count label
        self.component_count_label.setText(f"{len(self.component_widgets)} of 5")
        
        # Disable add button if limit reached
        if len(self.component_widgets) >= 5:
            self.add_component_btn.setEnabled(False)
            self.add_component_btn.setText("Maximum components reached")
        
        # Enable validation
        for widget in self.component_widgets:
            widget.name_input.textChanged.connect(self._validate_inputs)
        self._validate_inputs()
    
    def _save(self):
        """Save the equipment and components."""
        equipment_name = self.equipment_name_input.text().strip()
        serial_number = self.serial_input.text().strip() or None
        location = self.location_input.text().strip() or None
        components_data = []
        
        for widget in self.component_widgets:
            data = widget.get_data()
            if data:
                components_data.append(data)
        
        if equipment_name and components_data:
            if self.is_edit_mode:
                # Update existing machine
                old_name = self.machine_data['name']
                machine_id = self.machine_data['id']
                if not DatabaseManager.update_machine_details(machine_id, old_name, equipment_name, serial_number, location):
                    self._show_error_message("Failed to update equipment. Please try again.")
                    return
                
                existing_components = DatabaseManager.get_components_for_machine(machine_id)
                for index, comp_data in enumerate(components_data):
                    if index < len(existing_components):
                        component_id = existing_components[index]['id']
                        if not DatabaseManager.update_component(component_id, machine_id, equipment_name, comp_data):
                            self._show_error_message("Failed to update component. Please try again.")
                            return
                    else:
                        if not DatabaseManager.add_component(machine_id, equipment_name, comp_data):
                            self._show_error_message("Failed to add component. Please try again.")
                            return

                if len(existing_components) > len(components_data):
                    for component in existing_components[len(components_data):]:
                        if not DatabaseManager.delete_component(component['id']):
                            self._show_error_message("Failed to remove unused component. Please try again.")
                            return
                
                self.accept()
            else:
                # Add new machine to database
                machine_id = DatabaseManager.add_machine(equipment_name, serial_number, location, self.sheet_id)
                if machine_id is not None:
                    # Add components
                    for comp_data in components_data:
                        if not DatabaseManager.add_component(machine_id, equipment_name, comp_data):
                            self._show_error_message("Failed to add component. Please try again.")
                            return
                    
                    self.accept()
                else:
                    self._show_error_message("Failed to add equipment. Please try again.")
    
    def _show_error_message(self, message: str):
        """Show an error message with dark theme styling."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Error")
        msg_box.setText(message)
        msg_box.setIcon(QMessageBox.Icon.Warning)
        
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        msg_box.exec()

# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL CONFIG DIALOG
# ══════════════════════════════════════════════════════════════════════════════
class EmailConfigDialog(QDialog):
    """Dialog for configuring email settings."""
    
    def __init__(self, parent=None, sheet_id: Optional[int] = None):
        super().__init__(parent)
        self.sheet_id = sheet_id if sheet_id is not None else getattr(parent, "current_sheet_id", 1)
        self.sheet_data = DatabaseManager.get_sheet_by_id(self.sheet_id) or {"name": "Sheet", "email_recipients": []}
        self.config = EmailConfig.load_config()
        self._setup_ui()
    
    def _setup_ui(self):
        sheet_name = self.sheet_data.get("name", "Sheet")
        self.setWindowTitle(f"Email Configuration - {sheet_name}")
        self.setMinimumWidth(500)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(16)
        
        # Title
        title = QLabel(f"Email Configuration - {sheet_name}")
        title.setStyleSheet(f"""
            QLabel {{
                font-size: 20px;
                font-weight: 700;
                color: {Theme.TEXT_PRIMARY};
                border: none;
            }}
        """)
        layout.addWidget(title)
        
        # Enable checkbox
        self.enable_checkbox = QCheckBox("Enable Email Notifications")
        self.enable_checkbox.setChecked(self.config.get('enabled', False))
        self.enable_checkbox.setStyleSheet(f"""
            QCheckBox {{
                color: {Theme.TEXT_PRIMARY};
                font-size: 14px;
                spacing: 8px;
            }}
            QCheckBox::indicator {{
                width: 20px;
                height: 20px;
                border: 2px solid {Theme.BORDER};
                border-radius: 4px;
                background-color: {Theme.BG_INPUT};
            }}
            QCheckBox::indicator:checked {{
                background-color: {Theme.PRIMARY};
                border-color: {Theme.PRIMARY};
            }}
        """)
        layout.addWidget(self.enable_checkbox)
        
        # SMTP Host
        host_label = QLabel("SMTP Host:")
        host_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(host_label)
        
        self.host_input = StyledLineEdit("smtp.gmail.com")
        self.host_input.setText(self.config.get('smtp_host', 'smtp.gmail.com'))
        layout.addWidget(self.host_input)
        
        # SMTP Port
        port_label = QLabel("SMTP Port:")
        port_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(port_label)
        
        self.port_input = StyledSpinBox()
        self.port_input.setMaximum(65535)
        self.port_input.setValue(self.config.get('smtp_port', 587))
        layout.addWidget(self.port_input)
        
        # SMTP User
        user_label = QLabel("SMTP Username:")
        user_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(user_label)
        
        self.user_input = StyledLineEdit()
        self.user_input.setText(self.config.get('smtp_user', ''))
        self.user_input.setEchoMode(QLineEdit.EchoMode.Normal)
        layout.addWidget(self.user_input)
        
        # SMTP Password
        pass_label = QLabel("SMTP Password:")
        pass_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(pass_label)
        
        self.pass_input = StyledLineEdit()
        self.pass_input.setText(self.config.get('smtp_password', ''))
        self.pass_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.pass_input)
        
        # From Address
        from_label = QLabel("From Email:")
        from_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(from_label)
        
        self.from_input = StyledLineEdit()
        self.from_input.setText(self.config.get('from_addr', ''))
        layout.addWidget(self.from_input)
        
        # To Addresses
        to_label = QLabel("Recipients for this sheet (comma separated):")
        to_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; font-size: 14px; border: none; }}")
        layout.addWidget(to_label)
        
        self.to_input = StyledLineEdit()
        self.to_input.setText(', '.join(self.sheet_data.get('email_recipients', [])))
        layout.addWidget(self.to_input)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        test_btn = StyledButton("Send Test Email", primary=False)
        test_btn.clicked.connect(self._send_test_email)
        button_layout.addWidget(test_btn)
        
        cancel_btn = StyledButton("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        save_btn = StyledButton("Save", primary=True)
        save_btn.clicked.connect(self._save)
        button_layout.addWidget(save_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def _save(self):
        """Save the email configuration."""
        recipients = [addr.strip() for addr in self.to_input.text().split(',') if addr.strip()]
        config = {
            'enabled': self.enable_checkbox.isChecked(),
            'smtp_host': self.host_input.text().strip(),
            'smtp_port': self.port_input.value(),
            'smtp_user': self.user_input.text().strip(),
            'smtp_password': self.pass_input.text().strip(),
            'from_addr': self.from_input.text().strip(),
            'to_addrs': []
        }
        
        if EmailConfig.save_config(config) and DatabaseManager.update_sheet_recipients(self.sheet_id, recipients):
            self._show_success_message("Email configuration saved successfully!")
            self.accept()
        else:
            self._show_error_message("Failed to save email configuration.")
    
    def _send_test_email(self):
        """Send a test email using current configuration."""
        # Get current configuration from dialog inputs
        config = {
            'enabled': self.enable_checkbox.isChecked(),
            'smtp_host': self.host_input.text().strip(),
            'smtp_port': self.port_input.value(),
            'smtp_user': self.user_input.text().strip(),
            'smtp_password': self.pass_input.text().strip(),
            'from_addr': self.from_input.text().strip(),
            'to_addrs': [addr.strip() for addr in self.to_input.text().split(',') if addr.strip()]
        }
        
        # Validate required fields
        if not all([config['smtp_host'], config['smtp_user'], config['smtp_password'], 
                   config['from_addr'], config['to_addrs']]):
            self._show_error_message("Please fill in all email configuration fields before sending a test email.")
            return
        
        # Get components due for maintenance
        try:
            components = DatabaseManager.get_components_due_soon()
            if not components:
                self._show_error_message("No components are currently due for maintenance. Add some equipment with overdue or upcoming maintenance first.")
                return
            
            # Build email with test prefix
            subject, html = EmailConfig.build_alert_email(components)
            subject = "[TEST] " + subject  # Add test prefix
            
            # Send test email using dialog config (not saved config)
            success, error = EmailConfig.send_email(config, subject, html)
            
            if success:
                self._show_success_message("Test email sent successfully! Please check your inbox.")
            else:
                self._show_error_message(f"Failed to send test email: {error}")
        except Exception as e:
            self._show_error_message(f"Error sending test email: {str(e)}")
    
    def _show_success_message(self, message: str):
        """Show a success message with dark theme styling."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Success")
        msg_box.setText(message)
        msg_box.setIcon(QMessageBox.Icon.Information)
        
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        msg_box.exec()
    
    def _show_error_message(self, message: str):
        """Show an error message with dark theme styling."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Error")
        msg_box.setText(message)
        msg_box.setIcon(QMessageBox.Icon.Warning)
        
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        msg_box.exec()

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION WINDOW
# ══════════════════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    """Main application window."""
    
    def __init__(self):
        super().__init__()
        self.email_thread = None
        self.active_metric_filter = None
        self.search_query = ""
        self.current_sheet_id = 1
        self.sheets = []
        self._window_geometry_adjusting = False
        self._setup_ui()
        self._setup_tray()
        self._start_email_thread()
        
        # Set window flags to allow minimize but prevent resize/restore
        self.setWindowFlags(Qt.WindowType.Window | Qt.WindowType.WindowMinimizeButtonHint | Qt.WindowType.WindowCloseButtonHint)
        
        # Start maximized for the best user experience
        QTimer.singleShot(0, self._maximize_on_current_screen)
        
        # Force layout update after window is shown
        QTimer.singleShot(100, self._force_layout_update)

    def _maximize_on_current_screen(self):
        """Maximize the window on the current screen."""
        if self._window_geometry_adjusting:
            return

        screen = self.screen() or QApplication.primaryScreen()
        if not screen:
            return

        self._window_geometry_adjusting = True
        try:
            window_handle = self.windowHandle()
            if window_handle:
                window_handle.setScreen(screen)
            self.raise_()
            self.activateWindow()
            self.showMaximized()
        finally:
            self._window_geometry_adjusting = False
    
    def _force_layout_update(self):
        """Force layout update based on current window size."""
        window_width = self.width()
        
        # Determine column count based on window width (infinite rows)
        if window_width >= self.WIDTH_THRESHOLD:
            self.current_columns = 3
        else:
            self.current_columns = 2
        
        self._refresh_data()
    
    def _setup_ui(self):
        """Setup the main UI."""
        self.setWindowTitle("Equipment PM Tracker")
        # Remove minimum size since window is always maximized
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {Theme.BG_PRIMARY};
            }}
            /* Modern Scrollbar Styling */
            QScrollBar:vertical {{
                border: none;
                background: transparent;
                width: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.TEXT_MUTED};
                min-height: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar:horizontal {{
                border: none;
                background: transparent;
                height: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:horizontal {{
                background: {Theme.TEXT_MUTED};
                min-width: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Full-width header bar
        header = self._create_header()
        main_layout.addWidget(header)
        
        # Content container with margins for the rest of the app
        content_container = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(24, 24, 24, 24)
        content_layout.setSpacing(24)
        content_margins = content_layout.contentsMargins()

        # Sheet tabs
        sheet_row = QHBoxLayout()
        sheet_row.setContentsMargins(0, 0, 0, 0)
        sheet_row.setSpacing(10)

        self.sheet_tabs = QTabBar()
        self.sheet_tabs.setMovable(False)
        self.sheet_tabs.setExpanding(False)
        self.sheet_tabs.setDocumentMode(True)
        self.sheet_tabs.setDrawBase(False)
        self.sheet_tabs.setElideMode(Qt.TextElideMode.ElideRight)
        self.sheet_tabs.setStyleSheet(f"""
            QTabBar::tab {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_MUTED};
                border: 1px solid {Theme.BORDER};
                border-radius: 10px;
                padding: 8px 14px;
                margin-right: 6px;
                min-height: 18px;
                font-weight: 600;
            }}
            QTabBar::tab:selected {{
                background-color: {Theme.PRIMARY_LIGHT};
                color: {Theme.TEXT_PRIMARY};
                border-color: {Theme.PRIMARY};
            }}
            QTabBar::tab:hover {{
                color: {Theme.TEXT_PRIMARY};
                border-color: {Theme.PRIMARY};
            }}
        """)
        self.sheet_tabs.currentChanged.connect(self._on_sheet_changed)
        self.sheet_tabs.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.sheet_tabs.customContextMenuRequested.connect(self._show_sheet_tab_context_menu)

        sheet_tabs_widget = QWidget()
        sheet_tabs_widget.setFixedHeight(40)
        sheet_tabs_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        sheet_tabs_layout = QHBoxLayout(sheet_tabs_widget)
        sheet_tabs_layout.setContentsMargins(content_margins.left() - 15, 0, 0, 0)
        sheet_tabs_layout.setSpacing(0)
        sheet_tabs_layout.addWidget(self.sheet_tabs, 0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        self.add_sheet_btn = StyledButton("+ Sheet", primary=False)
        self.add_sheet_btn.setFixedHeight(36)
        self.add_sheet_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 10px;
                padding: 0px 14px;
                font-weight: 600;
                font-size: 12px;
            }}
            QPushButton:hover {{
                border-color: {Theme.PRIMARY};
                color: {Theme.PRIMARY};
            }}
        """)
        self.add_sheet_btn.clicked.connect(self._add_sheet)
        sheet_row.addWidget(self.add_sheet_btn, 0, Qt.AlignmentFlag.AlignRight)

        self.manage_sheets_btn = StyledButton("Manage", primary=False)
        self.manage_sheets_btn.setFixedHeight(36)
        self.manage_sheets_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 10px;
                padding: 0px 14px;
                font-weight: 600;
                font-size: 12px;
            }}
            QPushButton:hover {{
                border-color: {Theme.PRIMARY};
                color: {Theme.PRIMARY};
            }}
        """)
        self.manage_sheets_btn.clicked.connect(self._manage_sheets)
        sheet_row.addWidget(self.manage_sheets_btn)

        self.export_sheet_btn = StyledButton("Export", primary=False)
        self.export_sheet_btn.setFixedHeight(36)
        self.export_sheet_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 10px;
                padding: 0px 14px;
                font-weight: 600;
                font-size: 12px;
            }}
            QPushButton:hover {{
                border-color: {Theme.PRIMARY};
                color: {Theme.PRIMARY};
            }}
        """)
        self.export_sheet_btn.clicked.connect(self._export_current_sheet)
        sheet_row.addWidget(self.export_sheet_btn)

        sheet_widget = QWidget()
        sheet_widget.setLayout(sheet_row)
        sheet_widget.setFixedHeight(40)
        sheet_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        sheet_wrapper = QHBoxLayout()
        sheet_wrapper.setContentsMargins(0, 0, content_margins.right() - 8, 0)
        sheet_wrapper.setSpacing(0)
        sheet_wrapper.addStretch()
        sheet_wrapper.addWidget(sheet_widget, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)

        sheet_wrapper_widget = QWidget()
        sheet_wrapper_widget.setLayout(sheet_wrapper)

        top_band = QHBoxLayout()
        top_band.setContentsMargins(0, 0, 0, 0)
        top_band.setSpacing(12)
        top_band.addWidget(sheet_tabs_widget, 1)
        top_band.addWidget(sheet_wrapper_widget, 0, Qt.AlignmentFlag.AlignRight)

        top_band_widget = QWidget()
        top_band_widget.setLayout(top_band)
        content_layout.addWidget(top_band_widget)

        # Metrics panel
        metrics_panel = self._create_metrics_panel()
        content_layout.addWidget(metrics_panel)

        # Search bar
        search_row = QHBoxLayout()
        search_row.setContentsMargins(0, 0, 0, 0)
        search_row.setSpacing(8)

        search_label = QLabel("Search")
        search_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_PRIMARY}; font-weight: 600; border: none; font-size: 12px; }}")
        search_row.addWidget(search_label)

        self.search_input = StyledLineEdit("Search equipment, components, location, serial number")
        self.search_input.setMinimumHeight(34)
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {Theme.BG_INPUT};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 10px;
                padding: 6px 10px;
                font-size: 13px;
            }}
            QLineEdit:focus {{
                border-color: {Theme.BORDER_FOCUS};
            }}
        """)
        self.search_input.textChanged.connect(self._on_search_changed)
        search_row.addWidget(self.search_input, 1)

        clear_search_btn = StyledButton("Clear", primary=False)
        clear_search_btn.setFixedHeight(34)
        clear_search_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 10px;
                padding: 0px 14px;
                font-weight: 600;
                font-size: 12px;
            }}
            QPushButton:hover {{
                border-color: {Theme.PRIMARY};
                color: {Theme.PRIMARY};
            }}
        """)
        clear_search_btn.clicked.connect(lambda: self.search_input.setText(""))
        search_row.addWidget(clear_search_btn)

        search_widget = QWidget()
        search_widget.setLayout(search_row)
        search_widget.setFixedHeight(40)
        search_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        search_wrapper = QHBoxLayout()
        search_wrapper.setContentsMargins(0, 0, content_margins.right() - 8, 0)
        search_wrapper.setSpacing(0)
        search_wrapper.addStretch()
        search_wrapper.addWidget(search_widget, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)

        search_wrapper_widget = QWidget()
        search_wrapper_widget.setLayout(search_wrapper)
        content_layout.addWidget(search_wrapper_widget)
        
        # Equipment cards container
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setStyleSheet(f"""
            QScrollArea {{
                border: none;
                background-color: transparent;
            }}
            QScrollArea > QWidget > QWidget {{
                background-color: transparent;
            }}
            /* Modern Scrollbar Styling */
            QScrollBar:vertical {{
                border: none;
                background: transparent;
                width: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.TEXT_MUTED};
                min-height: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar:horizontal {{
                border: none;
                background: transparent;
                height: 8px;
                margin: 0px;
            }}
            QScrollBar::handle:horizontal {{
                background: {Theme.TEXT_MUTED};
                min-width: 30px;
                border-radius: 4px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {Theme.TEXT_PRIMARY};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)
        
        self.cards_container = QWidget()
        self.cards_container.setStyleSheet("background-color: transparent;")
        self.cards_layout = QGridLayout()
        self.cards_layout.setSpacing(16)
        self.cards_layout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        
        # Set uniform column stretches for consistent grid sizing
        for i in range(3):  # Max 3 columns
            self.cards_layout.setColumnStretch(i, 1)
        
        # Responsive grid configuration (infinite rows, responsive columns)
        self.WIDTH_THRESHOLD = 1200  # Threshold for switching between 2 and 3 columns
        self.current_columns = 3  # Start with maximized layout
        self.previous_columns = 3
        
        # Placeholders are handled dynamically in _refresh_data()
        
        self.cards_container.setLayout(self.cards_layout)
        self.scroll_area.setWidget(self.cards_container)
        content_layout.addWidget(self.scroll_area, 1)
        
        # Match metrics panel margins to cards layout + scrollbar to align columns perfectly
        margins = self.cards_layout.contentsMargins()
        # Scrollbar is 8px wide as per stylesheet
        metrics_panel.layout().setContentsMargins(margins.left(), 0, margins.right() + 8, 0)
        
        content_container.setLayout(content_layout)
        main_layout.addWidget(content_container)
        
        central_widget.setLayout(main_layout)
        self._load_sheet_tabs()
    
    def _create_header(self) -> QWidget:
        """Create the header section."""
        header = QFrame()
        header.setObjectName("headerBar")
        header.setStyleSheet(f"""
            QFrame#headerBar {{
                background-color: {Theme.BG_CARD};
                border-bottom: 1px solid {Theme.BORDER};
            }}
        """)
        
        layout = QHBoxLayout()
        layout.setContentsMargins(24, 12, 24, 12)
        layout.setSpacing(16)
        
        # Title section with logo
        title_section = QHBoxLayout()
        title_section.setSpacing(12)
        
        # Logo
        logo_path = os.path.join(get_base_path(), "logo.png")
        logo_label = QLabel()
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            # Scale logo to reasonable size
            scaled_pixmap = pixmap.scaled(80, 40, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
        else:
            # Fallback if logo doesn't exist - use minimalist industrial symbol
            logo_label.setText("⚙")
            logo_label.setStyleSheet(f"""
                QLabel {{
                    font-size: 28px;
                    border: none;
                    color: {Theme.PRIMARY};
                }}
            """)
        logo_label.setFixedSize(80, 40)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_section.addWidget(logo_label)
        
        # Title and subtitle
        title_layout = QVBoxLayout()
        title_layout.setSpacing(4)
        
        title = QLabel("Maintenance Tracker")
        title.setStyleSheet(f"""
            QLabel {{
                font-size: 18px;
                font-weight: 700;
                color: {Theme.TEXT_PRIMARY};
                border: none;
            }}
        """)
        title_layout.addWidget(title)
        
        subtitle = QLabel("Made by Sankar | v1.0")
        subtitle.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 12px; border: none; }}")
        title_layout.addWidget(subtitle)
        
        title_section.addLayout(title_layout)
        
        layout.addLayout(title_section)
        
        layout.addStretch()
        
        # Add equipment button
        add_btn = StyledButton("+ Add Equipment")
        add_btn.clicked.connect(self._add_equipment)
        layout.addWidget(add_btn)
        
        # Email config button
        email_btn = StyledButton("Email", primary=False)
        email_btn.clicked.connect(self._open_email_config)
        # Override styling for dark header background
        email_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {Theme.TEXT_PRIMARY};
                border: 2px solid {Theme.BORDER};
                border-radius: {Theme.CORNER_RADIUS}px;
                padding: 12px 24px;
                font-weight: 600;
                font-size: 14px;
            }}
            QPushButton:hover {{
                border-color: {Theme.PRIMARY};
                color: {Theme.PRIMARY};
            }}
            QPushButton:pressed {{
                background-color: {Theme.BG_MUTED};
            }}
        """)
        layout.addWidget(email_btn)
        
        header.setLayout(layout)
        return header
    
    def _create_metrics_panel(self) -> QWidget:
        """Create the metrics panel."""
        panel = QWidget()
        panel.setStyleSheet(f"""
            QWidget {{
                background-color: transparent;
            }}
        """)
        
        layout = QGridLayout()
        layout.setSpacing(16)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Set uniform column stretches to match equipment cards grid
        for i in range(3):  # 3 columns to match equipment grid
            layout.setColumnStretch(i, 1)
        
        # Total Equipment
        total_card = self._create_metric_card(
            "Total Equipment", 
            "0", 
            Theme.PRIMARY_LIGHT, 
            Theme.PRIMARY,
            "📦"
        )
        self.total_equipment_label = total_card.findChild(QLabel, "metric_value")
        layout.addWidget(total_card, 0, 0)
        
        # Needs Attention
        self.critical_card = self._create_metric_card(
            "Needs Attention", 
            "0", 
            Theme.RED_LIGHT, 
            Theme.RED,
            "⚠️"
        )
        self.critical_label = self.critical_card.findChild(QLabel, "metric_value")
        self.critical_card.clicked.connect(lambda: self._toggle_metric_filter("critical"))
        layout.addWidget(self.critical_card, 0, 1)
        
        # Upcoming
        self.warning_card = self._create_metric_card(
            "Upcoming", 
            "0", 
            Theme.YELLOW_LIGHT, 
            Theme.YELLOW,
            "📅"
        )
        self.warning_label = self.warning_card.findChild(QLabel, "metric_value")
        self.warning_card.clicked.connect(lambda: self._toggle_metric_filter("warning"))
        layout.addWidget(self.warning_card, 0, 2)
        
        panel.setLayout(layout)
        return panel
    
    def _create_metric_card(self, title: str, value: str, bg_color: str, text_color: str, icon: str = "📊") -> QWidget:
        """Create a single metric card."""
        card = MetricCard()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {Theme.BG_CARD};
                border: 1px solid {Theme.BORDER};
                border-radius: {Theme.CORNER_RADIUS}px;
            }}
            QFrame[active="true"] {{
                border: 2px solid {Theme.PRIMARY};
                background-color: rgba(59, 130, 246, 0.05);
            }}
        """)
        
        layout = QHBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Icon
        icon_label = QLabel(icon)
        icon_label.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        icon_label.setStyleSheet(f"""
            QLabel {{
                background-color: {bg_color};
                color: {text_color};
                padding: 12px;
                border: none;
                border-radius: 12px;
                font-size: 20px;
            }}
        """)
        layout.addWidget(icon_label)
        
        # Text
        text_layout = QVBoxLayout()
        text_layout.setSpacing(4)
        
        title_label = QLabel(title)
        title_label.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        title_label.setStyleSheet(f"QLabel {{ color: {Theme.TEXT_MUTED}; font-size: 14px; border: none; background-color: transparent; }}")
        text_layout.addWidget(title_label)

        value_label = QLabel(value)
        value_label.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        value_label.setObjectName("metric_value")
        value_label.setStyleSheet(f"""
            QLabel {{
                font-size: 24px;
                font-weight: 700;
                color: {text_color};
                border: none;
                background-color: transparent;
            }}
        """)
        text_layout.addWidget(value_label)
        
        layout.addLayout(text_layout)
        layout.addStretch()
        
        card.setLayout(layout)
        return card

    def _toggle_metric_filter(self, filter_name: str):
        """Toggle a metric filter and refresh the grid."""
        if self.active_metric_filter == filter_name:
            self.active_metric_filter = None
        else:
            self.active_metric_filter = filter_name

        self._refresh_data()

    def _on_search_changed(self, text: str):
        """Update the active search query and refresh the grid."""
        self.search_query = text.strip().lower()
        self._refresh_data()

    def _load_sheet_tabs(self):
        """Load all sheet tabs from the database."""
        self.sheets = DatabaseManager.get_sheets()
        self.sheet_tabs.blockSignals(True)
        while self.sheet_tabs.count():
            self.sheet_tabs.removeTab(0)

        for sheet in self.sheets:
            index = self.sheet_tabs.addTab(sheet.get('name', 'Sheet'))
            self.sheet_tabs.setTabData(index, sheet.get('id', 1))

        if not self.sheets:
            self.current_sheet_id = 1
        else:
            sheet_ids = [sheet.get('id', 1) for sheet in self.sheets]
            if self.current_sheet_id not in sheet_ids:
                self.current_sheet_id = sheet_ids[0]
            for index in range(self.sheet_tabs.count()):
                if self.sheet_tabs.tabData(index) == self.current_sheet_id:
                    self.sheet_tabs.setCurrentIndex(index)
                    break

        self.sheet_tabs.blockSignals(False)
        self._refresh_data()

    def _on_sheet_changed(self, index: int):
        """Switch the active sheet when a tab is selected."""
        if index < 0:
            return
        sheet_id = self.sheet_tabs.tabData(index)
        if not sheet_id:
            return
        if self.current_sheet_id == sheet_id:
            return
        self.current_sheet_id = sheet_id
        self.active_metric_filter = None
        self._refresh_data()

    def _add_sheet(self):
        """Create a new sheet/tab for isolated tracking."""
        name, accepted = QInputDialog.getText(
            self,
            "New Sheet",
            "Enter a name for this sheet:",
            QLineEdit.EchoMode.Normal
        )
        if not accepted:
            return

        sheet_name = name.strip()
        if not sheet_name:
            QMessageBox.warning(self, "Invalid Name", "Please enter a sheet name.")
            return

        if DatabaseManager.add_sheet(sheet_name) is None:
            QMessageBox.warning(self, "Duplicate Sheet", "That sheet name already exists or could not be created.")
            return

        self._load_sheet_tabs()
        for index in range(self.sheet_tabs.count()):
            if self.sheet_tabs.tabText(index) == sheet_name:
                self.sheet_tabs.setCurrentIndex(index)
                break

    def _show_sheet_tab_context_menu(self, pos):
        """Show rename/delete options for a sheet tab."""
        index = self.sheet_tabs.tabAt(pos)
        if index < 0:
            return

        self.sheet_tabs.setCurrentIndex(index)
        sheet_id = self.sheet_tabs.tabData(index)
        if not sheet_id:
            return

        sheet_name = self.sheet_tabs.tabText(index)

        menu = QMenu(self)
        rename_action = QAction("Rename", self)
        delete_action = QAction("Delete", self)
        rename_action.triggered.connect(lambda: self._rename_sheet_by_id(sheet_id, sheet_name))
        delete_action.triggered.connect(lambda: self._delete_sheet_by_id(sheet_id, sheet_name))
        menu.addAction(rename_action)
        menu.addAction(delete_action)
        menu.exec(self.sheet_tabs.mapToGlobal(pos))

    def _rename_sheet_by_id(self, sheet_id: int, current_name: str):
        """Rename a sheet using its id."""
        new_name, accepted = QInputDialog.getText(
            self,
            "Rename Sheet",
            "Enter a new sheet name:",
            QLineEdit.EchoMode.Normal,
            current_name
        )
        if not accepted:
            return

        new_name = new_name.strip()
        if not new_name:
            QMessageBox.warning(self, "Invalid Name", "Please enter a sheet name.")
            return

        if new_name == current_name:
            return

        if not DatabaseManager.update_sheet_name(sheet_id, new_name):
            QMessageBox.warning(self, "Rename Failed", "Could not rename the sheet.")
            return

        self._load_sheet_tabs()

    def _delete_sheet_by_id(self, sheet_id: int, sheet_name: str):
        """Delete a sheet using its id."""
        if sheet_id == 1:
            QMessageBox.information(self, "Default Sheet", "The Default sheet cannot be deleted.")
            return

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Delete Sheet")
        msg_box.setText(
            f"Delete '{sheet_name}' and all equipment, components, maintenance logs, and reminder history inside it?"
        )
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        if msg_box.exec() != QMessageBox.StandardButton.Yes:
            return

        if DatabaseManager.delete_sheet(sheet_id):
            self._load_sheet_tabs()
        else:
            QMessageBox.warning(self, "Delete Failed", "Could not delete the selected sheet.")

    def _manage_sheets(self):
        """Open the sheet management dialog."""
        dialog = SheetManagementDialog(self)
        dialog.exec()

    def _export_current_sheet(self):
        """Export the active sheet's data to a formatted Excel workbook."""
        sheet = DatabaseManager.get_sheet_by_id(self.current_sheet_id)
        sheet_name = sheet.get("name", "Sheet") if sheet else "Sheet"
        safe_sheet_name = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in sheet_name).strip() or "Sheet"

        default_path = os.path.join(get_base_path(), f"{safe_sheet_name}_PM_Export.xlsx")
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Sheet to Excel",
            default_path,
            "Excel Workbook (*.xlsx)"
        )
        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            machines = DatabaseManager.get_all_machines(self.current_sheet_id)
            components = []
            for machine in machines:
                for component in machine.get("components", []):
                    comp = dict(component)
                    comp["machine_name"] = machine.get("name", "")
                    comp["machine_serial_number"] = machine.get("serial_number", "")
                    comp["machine_location"] = machine.get("location", "")
                    components.append(comp)

            logs = DatabaseManager.get_maintenance_history(sheet_id=self.current_sheet_id)
            due_components = DatabaseManager.get_components_due_soon(self.current_sheet_id)

            wb = Workbook()
            ws_overview = wb.active
            ws_overview.title = "Overview"
            ws_equipment = wb.create_sheet("Equipment")
            ws_components = wb.create_sheet("Components")
            ws_logs = wb.create_sheet("Logs")

            title_fill = PatternFill("solid", fgColor="1E293B")
            header_fill = PatternFill("solid", fgColor="3B82F6")
            subheader_fill = PatternFill("solid", fgColor="334155")
            alt_fill = PatternFill("solid", fgColor="F8FAFC")
            white_font = Font(color="FFFFFF", bold=True)
            dark_font = Font(color="0F172A")
            header_border = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )
            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

            def style_title(ws, title_text):
                ws.merge_cells("A1:F1")
                cell = ws["A1"]
                cell.value = title_text
                cell.fill = title_fill
                cell.font = Font(color="FFFFFF", bold=True, size=14)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 24

            def style_meta_rows(ws, rows):
                for row_idx, label, value in rows:
                    ws[f"A{row_idx}"] = label
                    ws[f"A{row_idx}"].font = Font(bold=True, color="334155")
                    ws[f"B{row_idx}"] = value
                    ws[f"B{row_idx}"].font = Font(color="0F172A")
                    ws[f"A{row_idx}"].fill = subheader_fill
                    ws[f"B{row_idx}"].fill = PatternFill("solid", fgColor="FFFFFF")
                    ws[f"A{row_idx}"].border = thin_border
                    ws[f"B{row_idx}"].border = thin_border
                    ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
                    ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")

            def format_table(ws, headers, rows, start_row=1, freeze_row=None):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.border = header_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for r_idx, row in enumerate(rows, start_row + 1):
                    row_fill = alt_fill if (r_idx - start_row) % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.fill = row_fill
                        cell.font = dark_font
                        cell.border = thin_border
                        if c_idx in (1, 2, 3, 5, 6, 7, 8):
                            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                end_row = start_row + len(rows)
                ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{max(end_row, start_row)}"
                if freeze_row:
                    ws.freeze_panes = freeze_row

            def set_widths(ws, widths):
                for col_letter, width in widths.items():
                    ws.column_dimensions[col_letter].width = width

            # Overview
            style_title(ws_overview, f"Equipment PM Export - {sheet_name}")
            overview_rows = [
                (3, "Sheet Name", sheet_name),
                (4, "Exported On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                (5, "Equipment Count", len(machines)),
                (6, "Component Count", len(components)),
                (7, "Due Soon / Overdue", len(due_components)),
                (8, "Maintenance Logs", len(logs)),
            ]
            style_meta_rows(ws_overview, overview_rows)
            ws_overview["A10"] = "Quick Summary"
            ws_overview["A10"].font = Font(bold=True, color="FFFFFF")
            ws_overview["A10"].fill = header_fill
            ws_overview["A10"].alignment = Alignment(horizontal="left")
            ws_overview.merge_cells("A10:B10")
            ws_overview["A11"] = (
                "The workbook includes an equipment summary, a detailed component list, "
                "and the maintenance log history for this sheet."
            )
            ws_overview["A11"].alignment = Alignment(wrap_text=True, vertical="top")
            ws_overview["A11"].border = thin_border
            ws_overview.merge_cells("A11:F13")
            ws_overview.row_dimensions[11].height = 42
            set_widths(ws_overview, {"A": 22, "B": 40, "C": 16, "D": 16, "E": 16, "F": 16})

            # Equipment
            equipment_rows = []
            for machine in machines:
                machine_components = machine.get("components", [])
                soonest = min((c.get("days_remaining", 0) for c in machine_components), default=None)
                status = "No components"
                if machine_components:
                    if any(c.get("days_remaining", 0) <= 0 for c in machine_components):
                        status = "Attention"
                    elif any(0 < c.get("days_remaining", 0) <= c.get("alert_threshold_days", 5) for c in machine_components):
                        status = "Upcoming"
                    else:
                        status = "Healthy"

                component_lines = []
                for comp in machine_components:
                    component_lines.append(
                        f"{comp.get('component_name', 'Component')} - {comp.get('days_remaining', 0)}d left"
                    )

                equipment_rows.append([
                    machine.get("name", ""),
                    machine.get("serial_number", ""),
                    machine.get("location", ""),
                    "\n".join(component_lines) if component_lines else "No components",
                    soonest if soonest is not None else "",
                    status,
                ])

            format_table(
                ws_equipment,
                ["Equipment", "Serial Number", "Location", "Components", "Soonest Due", "Status"],
                equipment_rows,
                start_row=1,
                freeze_row="A2"
            )
            set_widths(ws_equipment, {"A": 26, "B": 18, "C": 18, "D": 48, "E": 14, "F": 16})

            # Components
            component_rows = []
            for comp in components:
                component_rows.append([
                    comp.get("machine_name", ""),
                    comp.get("machine_serial_number", ""),
                    comp.get("machine_location", ""),
                    comp.get("component_name", ""),
                    comp.get("pm_interval_days", ""),
                    comp.get("alert_threshold_days", ""),
                    comp.get("last_performed_date") or "",
                    comp.get("next_due_date") or "",
                    comp.get("days_remaining", ""),
                    comp.get("custom_start_date") or "",
                ])

            format_table(
                ws_components,
                [
                    "Equipment", "Serial Number", "Location", "Component", "Interval (Days)",
                    "Alert (Days)", "Last Performed", "Next Due", "Days Left", "Custom Start"
                ],
                component_rows,
                start_row=1,
                freeze_row="A2"
            )
            set_widths(ws_components, {
                "A": 24, "B": 16, "C": 18, "D": 22, "E": 14,
                "F": 12, "G": 14, "H": 14, "I": 12, "J": 14
            })

            # Logs
            log_rows = []
            for log in logs:
                log_rows.append([
                    log.get("maintenance_date", ""),
                    log.get("machine_name", ""),
                    log.get("machine_serial_number", ""),
                    log.get("component_name", ""),
                    log.get("notes") or "",
                ])

            format_table(
                ws_logs,
                ["Date", "Equipment", "Serial Number", "Component", "Notes"],
                log_rows,
                start_row=1,
                freeze_row="A2"
            )
            set_widths(ws_logs, {"A": 14, "B": 24, "C": 16, "D": 22, "E": 52})

            for ws in (ws_overview, ws_equipment, ws_components, ws_logs):
                ws.sheet_view.showGridLines = False

            wb.save(file_path)

            open_now = QMessageBox.question(
                self,
                "Export Complete",
                "Sheet exported successfully. Open the file now?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if open_now == QMessageBox.StandardButton.Yes:
                try:
                    os.startfile(file_path)
                except Exception as open_error:
                    QMessageBox.warning(self, "Open Failed", f"Could not open the exported file:\n{open_error}")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", f"Could not export the sheet:\n{e}")

    def _create_empty_slot(self) -> QWidget:
        """Create an empty placeholder for a grid slot."""
        placeholder = QWidget()
        placeholder.setFixedHeight(320)  # Match equipment card height
        placeholder.setStyleSheet(f"""
            QWidget {{
                background-color: transparent;
                border: 2px dashed {Theme.BORDER};
                border-radius: {Theme.CORNER_RADIUS}px;
            }}
        """)
        return placeholder
    
    def _initialize_grid_slots(self):
        """Initialize grid layout with placeholder slots for empty state."""
        # This is now handled in _refresh_data() for dynamic placeholder management
        pass
    
    def _setup_tray(self):
        """Setup system tray icon."""
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        
        # Create a small tray-and-arrow icon directly in code.
        pixmap = QPixmap(64, 64)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        accent = QColor(Theme.PRIMARY)
        base = QColor(Theme.BG_PRIMARY)
        white = QColor("#ffffff")

        # Tray base
        painter.setPen(QPen(accent, 3, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
        painter.setBrush(QBrush(base))
        painter.drawRoundedRect(12, 28, 40, 20, 6, 6)

        # Arrow
        painter.setPen(QPen(white, 4, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
        painter.drawLine(32, 16, 32, 34)
        arrow_head = QPolygon([
            QPoint(24, 26),
            QPoint(32, 36),
            QPoint(40, 26),
        ])
        painter.setBrush(QBrush(white))
        painter.drawPolygon(arrow_head)
        painter.end()

        icon = QIcon(pixmap)
        
        self.tray_icon = QSystemTrayIcon(icon)
        self.tray_icon.setToolTip("Production PM Tracker")
        
        # Create tray menu
        tray_menu = QMenu()
        
        show_action = QAction("Show", self)
        show_action.triggered.connect(self.show)
        tray_menu.addAction(show_action)
        
        quit_action = QAction("Quit", self)
        quit_action.triggered.connect(QApplication.instance().quit)
        tray_menu.addAction(quit_action)
        
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self._tray_activated)
        self.tray_icon.show()
    
    def _tray_activated(self, reason):
        """Handle tray icon activation."""
        if reason == QSystemTrayIcon.ActivationReason.DoubleClick:
            self.show()
    
    def _start_email_thread(self):
        """Start the email notification thread."""
        self.email_thread = EmailNotificationThread()
        self.email_thread.set_callback(self._email_notification_callback)
        self.email_thread.start()
    
    def _email_notification_callback(self, message: str):
        """Handle email notification callback."""
        pass
    
    def _add_equipment(self):
        """Open the add equipment dialog."""
        dialog = AddEquipmentDialog(sheet_id=self.current_sheet_id, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._refresh_data()
    
    def _edit_equipment(self, machine_data: Dict):
        """Open the edit equipment dialog."""
        dialog = AddEquipmentDialog(machine_data=machine_data, sheet_id=self.current_sheet_id, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._refresh_data()
    
    def _delete_equipment(self, machine_data: Dict):
        """Handle equipment deletion with confirmation."""
        machine_id = machine_data['id']
        machine_name = machine_data['name']
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Delete Equipment")
        msg_box.setText(f"Are you sure you want to delete '{machine_name}' and all its components?")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        # Style the box to match dark industrial theme
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        reply = msg_box.exec()
        
        if reply == QMessageBox.StandardButton.Yes:
            if DatabaseManager.delete_machine(machine_id):
                self._refresh_data()
    
    def _open_email_config(self):
        """Open the email configuration dialog."""
        dialog = EmailConfigDialog(self, sheet_id=self.current_sheet_id)
        dialog.exec()
    
    def _refresh_data(self):
        """Refresh all data from the database."""
        machines = DatabaseManager.get_all_machines(self.current_sheet_id)

        def _machine_due_sort_key(machine: Dict):
            component_days = [
                component.get('days_remaining', 0)
                for component in machine.get('components', [])
            ]
            # Machines with no components should drift to the bottom.
            soonest_due = min(component_days) if component_days else 10**9
            return (soonest_due, machine.get('name', '').lower())

        machines = sorted(machines, key=_machine_due_sort_key)

        def _machine_matches_filter(machine: Dict) -> bool:
            if not self.active_metric_filter:
                metric_match = True
            else:
                metric_match = False

                for component in machine.get('components', []):
                    days = component.get('days_remaining', 0)
                    alert_threshold = component.get('alert_threshold_days', 5)

                    if self.active_metric_filter == "critical" and days <= 0:
                        metric_match = True
                        break
                    if self.active_metric_filter == "warning" and 0 < days <= alert_threshold:
                        metric_match = True
                        break

            if not metric_match:
                return False

            if not self.search_query:
                return True

            search_terms = [self.search_query]
            machine_text = " ".join([
                machine.get('name', ''),
                machine.get('serial_number') or '',
                machine.get('location') or '',
                " ".join(
                    component.get('component_name', '')
                    for component in machine.get('components', [])
                    if component.get('component_name')
                ),
            ]).lower()
            return any(term in machine_text for term in search_terms)

        filtered_machines = [machine for machine in machines if _machine_matches_filter(machine)]
        
        # Clear all existing widgets from the layout using reversed loop
        for i in range(self.cards_layout.count() - 1, -1, -1):
            item = self.cards_layout.itemAt(i)
            if item:
                widget = item.widget()
                if widget:
                    widget.deleteLater()
        
        # Process events to ensure widgets are deleted before adding new ones
        QApplication.processEvents()
        
        # Update metrics
        total_equipment = len(filtered_machines)
        critical_count = 0
        warning_count = 0
        
        # Add cards for each machine with dynamic row calculation
        for i, machine in enumerate(filtered_machines):
            card = EquipmentCard(machine)
            card.reset_component.connect(self._reset_component)
            card.edit_requested.connect(self._edit_equipment)
            card.delete_requested.connect(self._delete_equipment)
            card.history_requested.connect(self._view_maintenance_history)
            card.refresh_needed.connect(self._refresh_data)
            
            # Count critical and warning components
            for component in machine.get('components', []):
                days = component.get('days_remaining', 0)
                alert_threshold = component.get('alert_threshold_days', 5)
                
                if days <= 0:
                    critical_count += 1
                elif days <= alert_threshold:
                    warning_count += 1
            
            # Dynamic row calculation for infinite scrolling
            row = i // self.current_columns
            col = i % self.current_columns
            
            # Add the card to the grid
            self.cards_layout.addWidget(card, row, col)
        
        # Add placeholder slots for empty positions (minimum 6 slots to show grid when empty)
        min_slots = 6  # Show at least a 2x3 grid when empty
        start_slot = len(filtered_machines)
        
        # Calculate how many placeholders we need
        # If we have equipment, fill the current row AND add one more full row
        # If no equipment, show minimum grid
        if filtered_machines:
            # Complete the current row
            current_row_items = len(filtered_machines) % self.current_columns
            if current_row_items == 0:
                # Row is already full, add one more full row
                placeholders_needed = self.current_columns
            else:
                # Fill current row and add one more full row
                placeholders_needed = (self.current_columns - current_row_items) + self.current_columns
        else:
            # Show minimum grid when empty
            placeholders_needed = min_slots
        
        for i in range(placeholders_needed):
            placeholder = self._create_empty_slot()
            slot_index = start_slot + i
            row = slot_index // self.current_columns
            col = slot_index % self.current_columns
            self.cards_layout.addWidget(placeholder, row, col)
        
        # Update metric labels
        self.total_equipment_label.setText(str(total_equipment))
        self.critical_label.setText(str(critical_count))
        self.warning_label.setText(str(warning_count))

        if hasattr(self, "critical_card"):
            self.critical_card.set_active(self.active_metric_filter == "critical")
        if hasattr(self, "warning_card"):
            self.warning_card.set_active(self.active_metric_filter == "warning")
        
        # Force UI update
        QApplication.processEvents()
    
    def resizeEvent(self, event):
        """Handle window resize events to force maximized state when restored."""
        if self._window_geometry_adjusting:
            super().resizeEvent(event)
            return

        # Only force back to maximized if window is visible but not maximized
        # This allows minimizing but forces maximized when restored
        if self.isVisible() and not self.isMaximized():
            self._maximize_on_current_screen()
            return  # Don't process the resize event
        
        super().resizeEvent(event)
        
        window_width = self.width()
        
        # Determine column count based on window width (infinite rows)
        if window_width >= self.WIDTH_THRESHOLD:
            target_columns = 3
        else:
            target_columns = 2
        
        # Only refresh if column count changed
        if target_columns != self.current_columns:
            self.current_columns = target_columns
            self._refresh_data()
        
        event.accept()
    
    def changeEvent(self, event):
        """Handle window state changes."""
        if event.type() == QEvent.Type.WindowStateChange and not self._window_geometry_adjusting:
            # If window is restored from minimized (but not maximized), force to maximized
            if self.isVisible() and not self.isMinimized() and not self.isMaximized():
                self._maximize_on_current_screen()
        super().changeEvent(event)
    
    def _reset_component(self, component_id: int):
        """Record a component's maintenance completion."""
        component = DatabaseManager.get_component_details(component_id)
        if not component:
            self._show_error_message("Could not load the selected component.")
            return

        dialog = RecordMaintenanceDialog(
            component.get('machine_name', 'Unknown Equipment'),
            component.get('component_name', 'Unknown Component'),
            self
        )

        if dialog.exec() == QDialog.DialogCode.Accepted:
            maintenance_date, notes = dialog.get_data()
            if DatabaseManager.record_component_maintenance(component_id, maintenance_date, notes):
                self._show_success_message("Maintenance recorded successfully!")
                self._refresh_data()
            else:
                self._show_error_message("Failed to record maintenance.")

    def _view_maintenance_history(self, machine_data: Dict):
        """Open the maintenance history dialog for a machine."""
        dialog = MaintenanceHistoryDialog(machine_data, sheet_id=machine_data.get('sheet_id', self.current_sheet_id), parent=self)
        dialog.exec()
    
    def _show_success_message(self, message: str):
        """Show a success message with dark theme styling."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Success")
        msg_box.setText(message)
        msg_box.setIcon(QMessageBox.Icon.Information)
        
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        msg_box.exec()
    
    def _show_error_message(self, message: str):
        """Show an error message with dark theme styling."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Error")
        msg_box.setText(message)
        msg_box.setIcon(QMessageBox.Icon.Warning)
        
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 6px 16px;
                min-width: 70px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        msg_box.exec()
    
    def closeEvent(self, event):
        """Handle window close event with custom options."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Exit Application")
        msg_box.setText("What would you like to do?")
        msg_box.setIcon(QMessageBox.Icon.Question)
        
        # Set minimum size to prevent button text cutoff
        msg_box.setMinimumWidth(400)
        msg_box.setMinimumHeight(150)
        
        # Create custom buttons
        minimize_btn = msg_box.addButton("Minimize to Tray", QMessageBox.ButtonRole.ActionRole)
        close_btn = msg_box.addButton("Close", QMessageBox.ButtonRole.DestructiveRole)
        cancel_btn = msg_box.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
        
        # Set default button to cancel
        msg_box.setDefaultButton(cancel_btn)
        
        # Style the box to match dark industrial theme
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: {Theme.BG_PRIMARY};
                color: {Theme.TEXT_PRIMARY};
            }}
            QLabel {{
                color: {Theme.TEXT_PRIMARY};
                background: transparent;
                border: none;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: {Theme.BG_CARD};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
                padding: 8px 16px;
                min-width: 120px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {Theme.BG_MUTED};
                border-color: {Theme.TEXT_MUTED};
            }}
        """)
        
        msg_box.exec()
        
        if msg_box.clickedButton() == minimize_btn:
            # Minimize to system tray
            self.hide()
            event.ignore()
        elif msg_box.clickedButton() == close_btn:
            # Close the application
            if self.email_thread:
                self.email_thread.stop()
            event.accept()
        else:
            # Cancel - keep application open
            event.ignore()

# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def main():
    """Main entry point for the application."""
    app = QApplication(sys.argv)
    
    # Initialize database
    init_database()
    
    # Create and show main window
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
